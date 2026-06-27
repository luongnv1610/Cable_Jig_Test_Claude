#!/usr/bin/env python3
"""
create_report_template.py
Tạo file Excel template cho cable test report.
"""
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import openpyxl
from datetime import datetime

# ── Color palette ─────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # Dark navy
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E75B6"   # Medium blue
C_SUBHDR_FG   = "FFFFFF"
C_PASS_BG     = "E2EFDA"   # Light green
C_PASS_FG     = "375623"
C_FAIL_BG     = "FCE4D6"   # Light red
C_FAIL_FG     = "843C0C"
C_OPEN_BG     = "FFF2CC"   # Light yellow
C_SHORT_BG    = "FCE4D6"   # Light red
C_MISWIRE_BG  = "DDEBF7"   # Light blue
C_ALT_ROW     = "F2F7FF"   # Alternate row
C_BORDER      = "8EA9C1"
C_INFO_BG     = "DEEAF1"
C_SECTION_BG  = "BDD7EE"

def thin_border(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style='medium', color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name='Arial', size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name='Arial', size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def create_template(path="cable_test_report_template.xlsx"):
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: TEST REPORT
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Test Report"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8   # top margin

    # ── Banner ──────────────────────────────────────────────────────────
    ws.merge_cells("B2:K2")
    c = ws["B2"]
    c.value = "CABLE CONTINUITY TEST REPORT"
    c.font = Font(name='Arial', size=18, bold=True, color=C_HEADER_FG)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:K3")
    c = ws["B3"]
    c.value = "NEU Co., Ltd."
    c.font = Font(name='Arial', size=10, color="AAAAAA")
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 8

    # ── Info block ──────────────────────────────────────────────────────
    def info_row(row, label, value_cell, value="", note=""):
        ws.row_dimensions[row].height = 22
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font  = body_font(bold=True)
        ws[f"B{row}"].fill  = fill(C_INFO_BG)
        ws[f"B{row}"].border = thin_border()
        ws[f"B{row}"].alignment = left()

        ws.merge_cells(f"{value_cell}{row}:E{row}")
        ws[f"{value_cell}{row}"].value = value
        ws[f"{value_cell}{row}"].font  = body_font()
        ws[f"{value_cell}{row}"].fill  = fill("FFFFFF")
        ws[f"{value_cell}{row}"].border = thin_border()
        ws[f"{value_cell}{row}"].alignment = left()
        return ws[f"{value_cell}{row}"]

    # Left column info
    for r in range(5, 13):
        ws.row_dimensions[r].height = 22

    labels_l = [
        (5,  "Cable Part Number",   "C", "{{PART_NUMBER}}"),
        (6,  "Cable Type",          "C", "{{CABLE_TYPE}}"),
        (7,  "Serial Number",       "C", "{{SERIAL_NUMBER}}"),
        (8,  "Cable Length",        "C", "{{CABLE_LENGTH}}"),
        (9,  "Test Date",           "C", "{{TEST_DATE}}"),
        (10, "Test Time",           "C", "{{TEST_TIME}}"),
        (11, "Operator",            "C", "{{OPERATOR}}"),
        (12, "Firmware Version",    "C", "{{FW_VERSION}}"),
    ]
    for row, label, vc, val in labels_l:
        info_row(row, label, vc, val)

    # Right column info
    labels_r = [
        (5,  "Total Pins Tested",   "H", "{{TOTAL_PINS}}"),
        (6,  "Expected Connections","H", "{{EXPECTED_CONN}}"),
        (7,  "Opens Found",         "H", "{{OPENS}}"),
        (8,  "Shorts Found",        "H", "{{SHORTS}}"),
        (9,  "Mis-wires Found",     "H", "{{MISWIRES}}"),
        (10, "Total Faults",        "H", "{{TOTAL_FAULTS}}"),
        (11, "Test Duration (s)",   "H", "{{DURATION}}"),
        (12, "COM Port",            "H", "{{COM_PORT}}"),
    ]
    for row, label, vc, val in labels_r:
        ws[f"G{row}"].value = label
        ws[f"G{row}"].font  = body_font(bold=True)
        ws[f"G{row}"].fill  = fill(C_INFO_BG)
        ws[f"G{row}"].border = thin_border()
        ws[f"G{row}"].alignment = left()
        ws.merge_cells(f"{vc}{row}:K{row}")
        ws[f"{vc}{row}"].value = val
        ws[f"{vc}{row}"].font  = body_font()
        ws[f"{vc}{row}"].fill  = fill("FFFFFF")
        ws[f"{vc}{row}"].border = thin_border()
        ws[f"{vc}{row}"].alignment = left()

    # ── RESULT badge ────────────────────────────────────────────────────
    ws.merge_cells("B13:K13")
    ws.row_dimensions[13].height = 40
    result_cell = ws["B13"]
    result_cell.value = "{{RESULT_TEXT}}"
    result_cell.font  = Font(name='Arial', size=20, bold=True, color="FFFFFF")
    result_cell.fill  = fill("375623")   # will be overwritten dynamically
    result_cell.alignment = center()
    result_cell.border = medium_border()

    ws.row_dimensions[14].height = 10

    # ── Fault table header ───────────────────────────────────────────────
    ws.row_dimensions[15].height = 28
    ws.merge_cells("B15:K15")
    c = ws["B15"]
    c.value = "FAULT DETAIL"
    c.font  = hdr_font(12)
    c.fill  = fill(C_SUBHDR_BG)
    c.alignment = center()
    c.border = thin_border()

    ws.row_dimensions[16].height = 24
    headers = ["#", "Fault Type", "J1 Pin", "J2 Pin",
               "Description", "Pass Direction", "Status"]
    cols    = ["B",  "C",          "E",      "F",
               "G",              "J",              "K"]
    spans   = [None, "C:D",        None,     None,
               "G:I",            None,             None]

    for i,(h,col,span) in enumerate(zip(headers,cols,spans)):
        if span:
            ws.merge_cells(f"{span[0]}16:{span[2]}16")
        c = ws[f"{col}16"]
        c.value = h
        c.font  = hdr_font(10)
        c.fill  = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = thin_border()

    # ── Placeholder fault rows (17-36 = 20 rows) ─────────────────────────
    fault_colors = {
        "OPEN_J1J2": C_OPEN_BG, "OPEN_J2J1": C_OPEN_BG,
        "SHORT": C_SHORT_BG, "MISWIRE": C_MISWIRE_BG
    }
    for r in range(17, 37):
        ws.row_dimensions[r].height = 20
        bg = C_ALT_ROW if (r % 2 == 0) else "FFFFFF"
        for col in ["B","C","D","E","F","G","H","I","J","K"]:
            ws[f"{col}{r}"].fill   = fill(bg)
            ws[f"{col}{r}"].border = thin_border()
            ws[f"{col}{r}"].font   = body_font()
            ws[f"{col}{r}"].alignment = center()

    # ── Footer ───────────────────────────────────────────────────────────
    ws.row_dimensions[37].height = 8
    ws.merge_cells("B38:K38")
    ws["B38"].value = "Generated by Cable Jig Tester PC Software  |  A1253 Series"
    ws["B38"].font  = Font(name='Arial', size=9, color="888888", italic=True)
    ws["B38"].alignment = center()

    # ── Column widths ────────────────────────────────────────────────────
    set_col_widths(ws, {
        "A":2,"B":20,"C":14,"D":10,"E":10,"F":10,
        "G":14,"H":14,"I":10,"J":16,"K":14
    })

    # ── Freeze panes ────────────────────────────────────────────────────
    ws.freeze_panes = "B17"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: SUMMARY (for multiple tests)
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test History")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:L1")
    ws2["A1"].value = "TEST HISTORY LOG"
    ws2["A1"].font  = Font(name='Arial', size=14, bold=True, color=C_HEADER_FG)
    ws2["A1"].fill  = fill(C_HEADER_BG)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 30

    hist_headers = ["#","Date","Time","Serial No.","Cable Type",
                    "Part Number","Pins Tested","Opens","Shorts",
                    "Mis-wires","Total Faults","Result"]
    ws2.row_dimensions[2].height = 24
    for i,h in enumerate(hist_headers,1):
        c = ws2.cell(2,i,h)
        c.font  = hdr_font(10)
        c.fill  = fill(C_SUBHDR_BG)
        c.alignment = center()
        c.border = thin_border()

    # Sample placeholder rows
    for r in range(3,8):
        ws2.row_dimensions[r].height = 20
        bg = C_ALT_ROW if r%2==0 else "FFFFFF"
        for ci in range(1,13):
            c = ws2.cell(r,ci)
            c.fill   = fill(bg)
            c.border = thin_border()
            c.font   = body_font()
            c.alignment = center()

    # Column widths
    hist_widths = [5,14,10,16,14,18,12,8,8,10,12,10]
    for i,w in enumerate(hist_widths,1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: PIN MAP reference
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Pin Map")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:S1")
    ws3["A1"].value = "CONNECTOR PIN MAP — A1253-B21 (116 pins: A×17 E×16 B×17 F×16 C×17 G×16 D×17)"
    ws3["A1"].font  = Font(name='Arial', size=13, bold=True, color=C_HEADER_FG)
    ws3["A1"].fill  = fill(C_HEADER_BG)
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    # Column headers: Col 1-17
    ws3.row_dimensions[2].height = 22
    ws3["A2"].value = "Row \\ Col"
    ws3["A2"].font  = hdr_font(10)
    ws3["A2"].fill  = fill(C_SUBHDR_BG)
    ws3["A2"].alignment = center()
    ws3["A2"].border = thin_border()
    for col in range(1,18):
        c = ws3.cell(2, col+1, col)
        c.font  = hdr_font(10)
        c.fill  = fill(C_SUBHDR_BG)
        c.alignment = center()
        c.border = thin_border()

    # Physical row order with correct pin counts
    rows_layout = [
        ("A", 17), ("E", 16), ("B", 17), ("F", 16),
        ("C", 17), ("G", 16), ("D", 17),
    ]
    # Build linear table
    pin_table = []
    for rname, rcount in rows_layout:
        for col in range(1, rcount+1):
            pin_table.append((rname, col))

    # Column header: max 17 cols
    max_cols = 17
    ws3.row_dimensions[2].height = 22
    ws3["A2"].value = "Row \\ Col"; ws3["A2"].font=hdr_font(10)
    ws3["A2"].fill=fill(C_SUBHDR_BG); ws3["A2"].alignment=center(); ws3["A2"].border=thin_border()
    for col in range(1, max_cols+1):
        c = ws3.cell(2, col+1, col)
        c.font=hdr_font(10); c.fill=fill(C_SUBHDR_BG)
        c.alignment=center(); c.border=thin_border()

    linear_counter = 0
    for ri, (row_label, row_count) in enumerate(rows_layout):
        r = ri + 3
        ws3.row_dimensions[r].height = 30
        rc = ws3.cell(r, 1, row_label)
        rc.font=hdr_font(10); rc.fill=fill(C_SECTION_BG)
        rc.alignment=center(); rc.border=thin_border()
        for col in range(1, max_cols+1):
            c = ws3.cell(r, col+1)
            if col <= row_count:
                linear = linear_counter + (col-1)
                c.value = f"{row_label}{col}\n({linear})"
                c.font = Font(name='Arial', size=8)
                bg = "FFFFFF" if (ri+col)%2==0 else C_ALT_ROW
                c.fill = fill(bg)
            else:
                c.fill = fill("D3D3D3")  # greyed out — pin doesn't exist
            c.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.border = thin_border()
        linear_counter += row_count

    for i in range(max_cols+1):
        ws3.column_dimensions[get_column_letter(i+1)].width = 9
    ws3.column_dimensions["A"].width = 11

    wb.save(path)
    print(f"Template saved: {path}")
    return path


if __name__ == "__main__":
    create_template("cable_test_report_template.xlsx")
