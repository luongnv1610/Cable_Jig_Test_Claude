#!/usr/bin/env python3
"""
Cable Jig Tester — PC Software
Giao tiếp với MCU qua USB-COM (FT232), hiển thị kết quả,
và generate report Excel từ template.

Requirements: pip install pyserial openpyxl
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
import time
import os
import sys
import json
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional
import shutil

# ── Try import openpyxl ───────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Constants ─────────────────────────────────────────────────────────────
APP_TITLE   = "Cable Jig Tester v4.5"
BAUD_RATE   = 115200
TIMEOUT_S   = 0.1
TEMPLATE_FILE = "cable_test_report_template.xlsx"
HISTORY_FILE  = "test_history.json"

# Colors
CLR_PASS    = "#2ECC71"
CLR_FAIL    = "#E74C3C"
CLR_WARN    = "#F39C12"
CLR_IDLE    = "#95A5A6"
CLR_BG      = "#1C2833"
CLR_PANEL   = "#2C3E50"
CLR_CARD    = "#34495E"
CLR_TEXT    = "#ECF0F1"
CLR_MUTED   = "#95A5A6"
CLR_ACCENT  = "#3498DB"
CLR_OPEN    = "#F39C12"
CLR_SHORT   = "#E74C3C"
CLR_MISWIRE = "#9B59B6"

TOTAL_PINS = 116

# ── GND common-ground pin sets ────────────────────────────────────────────
# Pins on the shared GND bus (verified from manual measurement).
GND_J1 = frozenset([
    "A2","A4","A6","A8","A10","A12","A14","A16",
    "E1","E3","E5","E7","E9","E11","E13","E15",
    "B1","B3","B5","B7","B9","B11","B13","B15","B17",
    "C2","C4","C6","C8","C10","C12","C14","C16",
    "G2","G4","G6","G8","G10","G12","G14","G16",
    "D1","D3","D5","D7","D9","D11","D13","D15","D17",
])
GND_J2 = frozenset([
    "A1","A3","A5","A7","A9","A11","A13","A15","A17",
    "E2","E4","E6","E8","E10","E12","E14","E16",
    "B2","B4","B6","B8","B10","B12","B14","B16",
    "C1","C3","C5","C7","C9","C11","C13","C15","C17",
    "G1","G3","G5","G7","G9","G11","G13","G15",
    "D2","D4","D6","D8","D10","D12","D14","D16",
])

# ── Expected signal connections (J1 → J2), verified from hardware ─────────
EXPECTED_J1_TO_J2: dict = {
    "A1":"D1",  "A3":"D3",  "A5":"D5",  "A7":"D7",  "A9":"D9",
    "A11":"D11","A13":"D13","A15":"D15","A17":"D17",
    "B2":"C2",  "B4":"C4",  "B6":"C6",  "B8":"C8",  "B10":"C10",
    "B12":"C12","B14":"C14","B16":"C16",
    "C1":"B1",  "C3":"B3",  "C5":"B5",  "C7":"B7",  "C9":"B9",
    "C11":"B11","C13":"B13","C15":"B15","C17":"B17",
    "D2":"A2",  "D4":"A4",  "D6":"A6",  "D8":"A8",  "D10":"A10",
    "D12":"A12","D14":"A14","D16":"A16",
    "E2":"G2",  "E4":"G4",  "E6":"G6",  "E8":"G8",  "E10":"G10",
    "E12":"G12","E14":"G14","E16":"G16",
    "F1":"F1",  "F2":"F2",  "F3":"F3",  "F4":"F4",  "F5":"F5",
    "F6":"F6",  "F7":"F7",  "F8":"F8",  "F9":"F9",  "F10":"F10",
    "F11":"F11","F12":"F12","F13":"F13","F14":"F14","F15":"F15","F16":"F16",
    "G1":"E1",  "G3":"E3",  "G5":"E5",  "G7":"E7",  "G9":"E9",
    "G11":"E11","G13":"E13","G15":"E15",
}
EXPECTED_J2_TO_J1: dict = {v: k for k, v in EXPECTED_J1_TO_J2.items()}



# ── Data classes ──────────────────────────────────────────────────────────
@dataclass
class ConnectionRecord:
    j1_pin: str
    j2_pin: str

@dataclass
class TestResult:
    timestamp: datetime               = field(default_factory=datetime.now)
    serial_number: str                = ""
    operator: str                     = ""
    fw_version: str                   = ""
    com_port: str                     = ""
    total_pins: int                   = 116
    connections: List[ConnectionRecord] = field(default_factory=list)
    duration_s: float                 = 0.0
    completed: bool                   = False
    raw_log: List[str]                = field(default_factory=list)


@dataclass
class JigPinResult:
    """Kết quả chẩn đoán phần cứng cho một J1 pin (Jig Self-Diagnostic)."""
    pin_name:       str
    status:         str        # "OK" hoặc "FAIL"
    conn_count:     int = 0    # số connection (chỉ dùng khi status="OK")
    mi:             int = -1   # MUX group index
    ch:             int = -1   # channel trong group
    lat:            int = -1   # ZA LAT readback (0=LOW/OK, 1=HIGH/FAIL)
    out0:           int = -1   # U19 OUT0 readback (-1 = không có/I2C lỗi)
    out1:           int = -1   # U19 OUT1 readback (-1 = không có/I2C lỗi)
    i2c_err:        bool = False  # firmware báo I2C error khi đọc thanh ghi
    root_cause:     str = ""   # nguyên nhân phân tích
    recommendation: str = ""   # hành động khắc phục

# ── Excel Report Generator ────────────────────────────────────────────────
class ReportGenerator:

    def thin(self, color="8EA9C1"):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(self, color):
        return PatternFill("solid", fgColor=color)

    def generate(self, result: TestResult, template_path: str, out_path: str) -> str:
        if not EXCEL_OK:
            raise RuntimeError("openpyxl not installed")

        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
        else:
            from create_template import create_template
            create_template(template_path)
            wb = openpyxl.load_workbook(template_path)

        ws = wb["Test Report"]
        n_conn = len(result.connections)

        # ── Info cells ───────────────────────────────────────────────────
        n_expected = len(EXPECTED_J1_TO_J2) + len(GND_J1)  # 66 signal + 50 GND = 116
        info_map = {
            "C5":  "—",
            "C6":  "Cable Continuity Scanner",
            "C7":  result.serial_number,
            "C8":  "—",
            "C9":  result.timestamp.strftime("%Y-%m-%d"),
            "C10": result.timestamp.strftime("%H:%M:%S"),
            "C11": result.operator or "—",
            "C12": result.fw_version or "—",
            "H5":  str(result.total_pins),
            "H6":  str(n_conn),
            "H7":  "0",
            "H8":  "0",
            "H9":  "—",
            "H10": "—",
            "H11": f"{result.duration_s:.1f}",
            "H12": result.com_port,
        }
        for ref, val in info_map.items():
            ws[ref] = val
            ws[ref].font = Font(name='Arial', size=10)

        # ── Separate signal vs GND connections ───────────────────────────
        # Dedup signal: for each signal J1 pin, take first non-GND J2 hit.
        # Also detect SHORT: signal J1 pin that also has GND_J2 hits = shorted to GND.
        from collections import defaultdict as _dd
        _j1_gnd  = _dd(list)   # signal J1 pin -> list of its GND_J2 hits
        _j1_sig  = _dd(list)   # signal J1 pin -> list of its signal J2 hits
        gnd_conns: list = []

        for c in result.connections:
            if c.j1_pin in GND_J1 or c.j2_pin in GND_J2:
                gnd_conns.append(c)
                if c.j1_pin not in GND_J1:          # signal J1 pulling GND bus
                    _j1_gnd[c.j1_pin].append(c.j2_pin)
            else:
                _j1_sig[c.j1_pin].append(c.j2_pin)  # real signal hit

        # Build deduped signal_conns (first hit per J1 pin)
        _seen_j1: set = set()
        signal_conns: list = []
        for c in result.connections:
            if c.j1_pin in GND_J1 or c.j2_pin in GND_J2:
                continue
            if c.j1_pin not in _seen_j1:
                _seen_j1.add(c.j1_pin)
                signal_conns.append(c)

        n_signal = len(signal_conns)
        n_gnd    = len(gnd_conns)

        # SHORT faults: signal J1 pins that also pulled GND bus
        short_faults = sorted(
            [j1 for j1 in _j1_gnd if _j1_gnd[j1]],
            key=lambda x: (x[0], int(x[1:]))
        )

        # Validate signal pairs for PASS/FAIL verdict
        detected = {c.j1_pin: c.j2_pin for c in signal_conns}
        missing_sig  = [j1 for j1, ej2 in EXPECTED_J1_TO_J2.items()
                        if detected.get(j1) is None]
        miswire_sig  = [f"J1.{j1}→got J2.{detected[j1]} exp J2.{ej2}"
                        for j1, ej2 in EXPECTED_J1_TO_J2.items()
                        if detected.get(j1) is not None and detected[j1] != ej2]
        verdict     = "PASS" if not missing_sig and not miswire_sig and not short_faults else "FAIL"
        verdict_clr = "1E8449" if verdict == "PASS" else "922B21"

        # Update info cells with verdict
        info_map["H7"]  = str(len(missing_sig))
        info_map["H8"]  = str(len(miswire_sig) + len(short_faults))
        info_map["H9"]  = verdict
        for ref in ("H7","H8","H9"):
            ws[ref] = info_map[ref]
            ws[ref].font = Font(name='Arial', size=10)
        ws["H9"].fill = self.fill(verdict_clr)
        ws["H9"].font = Font(name='Arial', size=10, bold=True, color="FFFFFF")

        # ── Result badge (matches screenshot) ────────────────────────────
        result_cell = ws["B13"]
        issues = len(missing_sig) + len(miswire_sig) + len(short_faults)
        if verdict == "PASS":
            result_cell.value = (
                f"✓  PASS — {n_signal} signal pair(s) verified"
                f"  |  GND bus: {n_gnd} connection(s)"
            )
        else:
            result_cell.value = (
                f"✗  FAIL — {issues} issue(s)"
                f"  |  {n_signal} signal detected"
                f"  |  GND bus: {n_gnd} conn"
            )
        result_cell.fill = self.fill(verdict_clr)
        result_cell.font = Font(name='Arial', size=15, bold=True, color="FFFFFF")

        # ── Connection rows ───────────────────────────────────────────────
        # Layout per image:
        #   Row 16 (header): # | J1 Pin | Pass Direction | J2 Pin
        #   Row 17+: GND block (merged, lists all J1 GND + J2 GND pins)
        #            then signal rows
        START_ROW = 17
        data_rows = max(n_signal + 20, 40)

        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= START_ROW - 1:
                ws.unmerge_cells(str(rng))
        for r in range(START_ROW - 1, START_ROW + data_rows):
            for col in ["B","C","D","E","F","G","H","I","J","K"]:
                ws[f"{col}{r}"].value = None

        # ── Column header row ─────────────────────────────────────────────
        HDR_ROW = START_ROW - 1
        ws.row_dimensions[HDR_ROW].height = 22
        ws.merge_cells(f"D{HDR_ROW}:J{HDR_ROW}")
        for ref, val, align in [
            (f"B{HDR_ROW}", "#",              "center"),
            (f"C{HDR_ROW}", "J1 Pin",         "center"),
            (f"D{HDR_ROW}", "Pass Direction", "center"),
            (f"K{HDR_ROW}", "J2 Pin",         "center"),
        ]:
            c = ws[ref]
            c.value = val
            c.font  = Font(name='Arial', size=10, bold=True, color="FFFFFF")
            c.fill  = self.fill("2E4F6F")
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.border = self.thin("1A3A5C")

        if not n_conn:
            ws.row_dimensions[START_ROW].height = 24
            ws.merge_cells(f"B{START_ROW}:K{START_ROW}")
            c = ws[f"B{START_ROW}"]
            c.value = "No connections detected"
            c.font  = Font(name='Arial', size=11, bold=True, color="843C0C")
            c.fill  = self.fill("FCE4D6")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = self.thin()
        else:
            row_idx = 0

            # ── GND block: one merged cell group ─────────────────────────
            # Collect all unique GND J1 and J2 pins from detected connections
            gnd_j1_sorted = sorted(
                {c.j1_pin for c in gnd_conns if c.j1_pin in GND_J1},
                key=lambda x: (x[0], int(x[1:]))
            )
            gnd_j2_sorted = sorted(
                {c.j2_pin for c in gnd_conns if c.j2_pin in GND_J2},
                key=lambda x: (x[0], int(x[1:]))
            )

            # Build display strings — each connector row (A,B,C...) on its own line
            def fmt_pin_list(pins, prefix):
                from itertools import groupby
                lines_out = []
                for row_letter, group in groupby(pins, key=lambda x: x[0]):
                    row_pins = list(group)
                    lines_out.append(", ".join(f"{prefix}.{p}" for p in row_pins))
                return "\n".join(lines_out)

            j1_text = fmt_pin_list(gnd_j1_sorted, "J1")
            j2_text = fmt_pin_list(gnd_j2_sorted, "J2")
            combined_text = j1_text + ("\n" if j1_text and j2_text else "") + j2_text

            # Calculate number of rows needed (one per text line, min 2)
            n_lines = combined_text.count("\n") + 1
            gnd_rows = max(n_lines + 1, 3)   # +1 for "GROUP" label line

            GND_START = START_ROW + row_idx
            GND_END   = GND_START + gnd_rows - 1

            # Row number cell (spans all GND rows)
            ws.merge_cells(f"B{GND_START}:B{GND_END}")
            c = ws[f"B{GND_START}"]
            c.value     = 1
            c.font      = Font(name='Arial', size=10, bold=True, color="7D6608")
            c.fill      = self.fill("FFFDE7")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = self.thin("C9A227")

            # J1 pin cell (spans all GND rows, left column)
            ws.merge_cells(f"C{GND_START}:C{GND_END}")
            c = ws[f"C{GND_START}"]
            c.value     = ""   # blank — pins listed inside D:J merged block
            c.fill      = self.fill("FFFDE7")
            c.border    = self.thin("C9A227")

            # Middle merged block D:J spanning all GND rows
            ws.merge_cells(f"D{GND_START}:J{GND_END}")
            c = ws[f"D{GND_START}"]
            c.value     = "GROUP\n" + combined_text
            c.font      = Font(name='Arial', size=10, color="7D6608")
            c.fill      = self.fill("FFFDE7")
            c.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
            c.border    = self.thin("C9A227")

            # J2 pin cell (spans all GND rows, right column)
            ws.merge_cells(f"K{GND_START}:K{GND_END}")
            c = ws[f"K{GND_START}"]
            c.value     = ""   # blank — pins listed inside D:J merged block
            c.fill      = self.fill("FFFDE7")
            c.border    = self.thin("C9A227")

            # Set row heights for GND block
            row_height = max(15 * gnd_rows, 80)
            for r in range(GND_START, GND_END + 1):
                ws.row_dimensions[r].height = row_height / gnd_rows

            row_idx += gnd_rows

            # ── Signal rows ───────────────────────────────────────────────
            for conn in signal_conns:
                r  = START_ROW + row_idx
                bg = "F2F7FF" if row_idx % 2 == 0 else "FFFFFF"
                exp_j2 = EXPECTED_J1_TO_J2.get(conn.j1_pin)
                if exp_j2 is None:
                    row_clr = "FDEDEC"
                    j2_val  = f"J2.{conn.j2_pin}  ⚠"
                elif conn.j2_pin == exp_j2:
                    row_clr = bg
                    j2_val  = f"J2.{conn.j2_pin}  ✓"
                else:
                    row_clr = "F3E5F5"
                    j2_val  = f"J2.{conn.j2_pin}  ✗ (exp J2.{exp_j2})"

                ws.row_dimensions[r].height = 18
                ws.merge_cells(f"D{r}:J{r}")
                for ref, val, align in [
                    (f"B{r}", row_idx + 1,          "center"),
                    (f"C{r}", f"J1.{conn.j1_pin}",  "center"),
                    (f"D{r}", "↔",                   "center"),
                    (f"K{r}", j2_val,                "left"),
                ]:
                    c = ws[ref]
                    c.value = val
                    c.font  = Font(name='Arial', size=10)
                    c.fill  = self.fill(row_clr)
                    c.alignment = Alignment(horizontal=align, vertical='center')
                    c.border = self.thin()
                row_idx += 1

            # ── Missing signal rows ───────────────────────────────────────
            for j1 in missing_sig:
                r = START_ROW + row_idx
                ws.row_dimensions[r].height = 18
                ws.merge_cells(f"D{r}:J{r}")
                for ref, val, align in [
                    (f"B{r}", row_idx + 1,                                    "center"),
                    (f"C{r}", f"J1.{j1}",                                     "center"),
                    (f"D{r}", "↔",                                             "center"),
                    (f"K{r}", f"MISSING  (exp J2.{EXPECTED_J1_TO_J2[j1]})",   "left"),
                ]:
                    c = ws[ref]
                    c.value = val
                    c.font  = Font(name='Arial', size=10, bold=(ref.startswith("K")))
                    c.fill  = self.fill("FDEDEC")
                    c.alignment = Alignment(horizontal=align, vertical='center')
                    c.border = self.thin("E57373")
                row_idx += 1

            # ── SHORT fault rows ──────────────────────────────────────────
            for j1 in short_faults:
                r = START_ROW + row_idx
                ws.row_dimensions[r].height = 18
                ws.merge_cells(f"D{r}:J{r}")
                for ref, val, align in [
                    (f"B{r}", row_idx + 1,                                         "center"),
                    (f"C{r}", f"J1.{j1}",                                          "center"),
                    (f"D{r}", "⚡ SHORT",                                           "center"),
                    (f"K{r}", f"SHORT to GND bus  (J1.{j1} bridged to GND)",       "left"),
                ]:
                    c = ws[ref]
                    c.value = val
                    c.font  = Font(name='Arial', size=10,
                                   bold=True, color=("FFFFFF" if ref.startswith("D") else "C0392B"))
                    c.fill  = self.fill("C0392B" if ref.startswith("D") else "FDEDEC")
                    c.alignment = Alignment(horizontal=align, vertical='center')
                    c.border = self.thin("C0392B")
                row_idx += 1

        end_r = START_ROW + max(row_idx if n_conn else 1, 1) + 2
        ws.merge_cells(f"B{end_r}:K{end_r}")
        ws[f"B{end_r}"].value = (
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            "  |  Cable Jig Tester PC Software"
        )
        ws[f"B{end_r}"].font      = Font(name='Arial', size=9, color="888888", italic=True)
        ws[f"B{end_r}"].alignment = Alignment(horizontal='center', vertical='center')

        # ── Test History sheet ────────────────────────────────────────────
        try:
            ws2 = wb["Test History"]
            next_row = 3
            for row in ws2.iter_rows(min_row=3):
                if row[0].value is not None:
                    next_row = row[0].row + 1
                else:
                    break
            ws2.row_dimensions[next_row].height = 20
            bg = "F2F7FF" if next_row % 2 == 0 else "FFFFFF"
            hist_data = [
                next_row - 2,
                result.timestamp.strftime("%Y-%m-%d"),
                result.timestamp.strftime("%H:%M:%S"),
                result.serial_number,
                "—",
                "—",
                result.total_pins,
                n_conn,
                "—", "—", "—",
                "DONE",
            ]
            for ci, val in enumerate(hist_data, 1):
                c = ws2.cell(next_row, ci, val)
                c.font  = Font(name='Arial', size=10)
                c.fill  = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = self.thin()
        except Exception:
            pass

        # ── Connection Map sheet ──────────────────────────────────────────
        try:
            if "Connection Map" in wb.sheetnames:
                del wb["Connection Map"]
            wm = wb.create_sheet("Connection Map")
            wm.sheet_view.showGridLines = False

            wm.merge_cells("A1:E1")
            wm["A1"].value = (
                f"Connection Map — S/N: {result.serial_number}"
                f"  |  {n_signal} signal pair(s)  +  {len(gnd_j1_sorted)} GND pin(s)"
            )
            wm["A1"].font  = Font(name='Arial', size=13, bold=True, color="FFFFFF")
            wm["A1"].fill  = self.fill(verdict_clr)
            wm["A1"].alignment = Alignment(horizontal='center', vertical='center')
            wm.row_dimensions[1].height = 28

            for col, hdr, w in zip("ABCDE",
                                   ["#", "J1 Pin", "Pass Direction", "J2 Pin", "Status"],
                                   [5, 14, 40, 14, 14]):
                c = wm[f"{col}2"]
                c.value = hdr
                c.font  = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                c.fill  = self.fill("2E4F6F")
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = self.thin()
                wm.column_dimensions[col].width = w
            wm.row_dimensions[2].height = 22

            cm_row = 3
            # GND block — one merged cell group in column C (Pass Direction)
            n_gnd_lines = combined_text.count("\n") + 1
            gnd_blk = max(n_gnd_lines + 1, 3)
            GND_END_CM = cm_row + gnd_blk - 1

            wm.merge_cells(f"A{cm_row}:A{GND_END_CM}")
            c = wm[f"A{cm_row}"]
            c.value = 1; c.font = Font(name='Arial',size=10,bold=True,color="7D6608")
            c.fill = self.fill("FFFDE7")
            c.alignment = Alignment(horizontal='center',vertical='center')
            c.border = self.thin("C9A227")

            wm.merge_cells(f"B{cm_row}:B{GND_END_CM}")
            c = wm[f"B{cm_row}"]
            c.value = ""; c.fill = self.fill("FFFDE7"); c.border = self.thin("C9A227")

            wm.merge_cells(f"C{cm_row}:C{GND_END_CM}")
            c = wm[f"C{cm_row}"]
            c.value = "GROUP\n" + combined_text
            c.font  = Font(name='Arial', size=10, color="7D6608")
            c.fill  = self.fill("FFFDE7")
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = self.thin("C9A227")

            wm.merge_cells(f"D{cm_row}:D{GND_END_CM}")
            c = wm[f"D{cm_row}"]
            c.value = ""; c.fill = self.fill("FFFDE7"); c.border = self.thin("C9A227")

            wm.merge_cells(f"E{cm_row}:E{GND_END_CM}")
            c = wm[f"E{cm_row}"]
            c.value = "GND"; c.font = Font(name='Arial',size=10,bold=True,color="7D6608")
            c.fill = self.fill("FFFDE7")
            c.alignment = Alignment(horizontal='center',vertical='center')
            c.border = self.thin("C9A227")

            rh = max(15 * gnd_blk, 80) / gnd_blk
            for r in range(cm_row, GND_END_CM + 1):
                wm.row_dimensions[r].height = rh
            cm_row = GND_END_CM + 1

            # Signal rows
            for idx, conn in enumerate(signal_conns, 2):
                bg = "F2F7FF" if idx % 2 == 0 else "FFFFFF"
                exp_j2 = EXPECTED_J1_TO_J2.get(conn.j1_pin)
                if exp_j2 is None:
                    status, bg = "⚠ Unexpected", "FDEDEC"
                elif conn.j2_pin == exp_j2:
                    status = "✓ OK"
                else:
                    status, bg = f"✗ exp J2.{exp_j2}", "F3E5F5"
                wm.row_dimensions[cm_row].height = 18
                for col, val, align in zip("ABCDE",
                    [idx, f"J1.{conn.j1_pin}", "↔", f"J2.{conn.j2_pin}", status],
                    ["center","center","center","center","left"]):
                    c = wm[f"{col}{cm_row}"]
                    c.value = val
                    c.font  = Font(name='Arial', size=10)
                    c.fill  = self.fill(bg)
                    c.alignment = Alignment(horizontal=align, vertical='center')
                    c.border = self.thin()
                cm_row += 1

            # Missing rows
            for j1 in missing_sig:
                wm.row_dimensions[cm_row].height = 18
                for col, val, align in zip("ABCDE",
                    [cm_row-2, f"J1.{j1}", "↔",
                     f"J2.{EXPECTED_J1_TO_J2[j1]}", "✗ MISSING"],
                    ["center","center","center","center","left"]):
                    c = wm[f"{col}{cm_row}"]
                    c.value = val
                    c.font  = Font(name='Arial', size=10,
                                   bold=(col=="E"), color=("C0392B" if col=="E" else "000000"))
                    c.fill  = self.fill("FDEDEC")
                    c.alignment = Alignment(horizontal=align, vertical='center')
                    c.border = self.thin("E57373")
                cm_row += 1

        except Exception:
            pass

        wb.save(out_path)
        return out_path


# ── Serial Reader Thread ──────────────────────────────────────────────────
class SerialReader(threading.Thread):
    def __init__(self, port, baud, rx_queue, stop_event):
        super().__init__(daemon=True)
        self.port = port
        self.baud = baud
        self.q    = rx_queue
        self.stop = stop_event
        self.ser  = None

    def run(self):
        try:
            self.ser = serial.Serial(self.port, self.baud,
                                     timeout=TIMEOUT_S,
                                     write_timeout=1.0)
            self.q.put(("CONNECTED", self.port))
            buf = ""
            while not self.stop.is_set():
                try:
                    chunk = self.ser.read(64).decode("ascii", errors="replace")
                    if chunk:
                        buf += chunk
                        while "\n" in buf:
                            line, buf = buf.split("\n", 1)
                            line = line.strip()
                            if line:
                                self.q.put(("LINE", line))
                except serial.SerialTimeoutException:
                    pass
                except Exception as e:
                    self.q.put(("ERROR", str(e)))
                    break
        except Exception as e:
            self.q.put(("ERROR", str(e)))
        finally:
            if self.ser and self.ser.is_open:
                self.ser.close()
            self.q.put(("DISCONNECTED", self.port))

    def send(self, cmd: str):
        if self.ser and self.ser.is_open:
            self.ser.write((cmd + "\r\n").encode())


# ── Main Application ──────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x780")
        self.minsize(900, 640)
        self.configure(bg=CLR_BG)
        self.resizable(True, True)

        # State
        self.serial_reader: Optional[SerialReader] = None
        self.serial_stop   = threading.Event()
        self.rx_queue      = queue.Queue()
        self.current_result: Optional[TestResult] = None
        self.test_start_time: Optional[float] = None
        self.in_test       = False
        self.test_stopped  = False          # user pressed STOP
        self.test_timeout_s = 120           # auto-stop after 120s no response
        self.history: List[TestResult] = []
        self.operator_name = ""

        # Jig self-diagnostic state — PASS 1 (J1 side)
        self._jig_fail_pins:     List[str]           = []
        self._jig_results_all:   List[JigPinResult]  = []
        self._sweep_data:        dict                 = {}   # {mi_str: set(ch_int)}
        self._sweep_hit_count:   int                 = 0
        self._sweep_mi:          Optional[str]       = None
        self._jig_ok_count:      int                 = 0
        self._jig_diag_ts:       Optional[datetime]  = None
        self._jig_diag_com:      str                 = ""
        self._jig_diag_fw:       str                 = ""
        # Jig self-diagnostic state — PASS 2 (J2 side)
        self._jig_j2_fail_pins:     List[str]           = []
        self._jig_j2_results_all:   List[JigPinResult]  = []
        self._sweep_j2_data:        dict                 = {}   # {mi_str: set(ch_int)}
        self._sweep_j2_hit_count:   int                 = 0
        self._sweep_j2_mi:          Optional[str]       = None
        self._jig_j2_ok_count:      int                 = 0
        # F-row even-pin debug
        self._fdiag_rows:           list                = []
        # J2 MUX-B group-0 sweep
        self._j2g0_rows:            list                = []
        # GND grouping state
        self._gnd_tree_id:          Optional[str]       = None
        self._gnd_count:            int                 = 0

        self._build_ui()
        self._load_history()
        self._refresh_ports()
        self._poll_queue()

    # ── UI Build ──────────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_topbar()
        self._build_main()
        self._build_statusbar()

    def _build_topbar(self):
        top = tk.Frame(self, bg=CLR_PANEL, height=60)
        top.pack(fill="x", side="top")
        top.pack_propagate(False)

        tk.Label(top, text="⚡ " + APP_TITLE,
                 font=("Arial",16,"bold"), fg=CLR_TEXT, bg=CLR_PANEL
                 ).pack(side="left", padx=20, pady=10)

        # About button
        tk.Button(top, text="ℹ  About",
                  font=("Arial", 9), fg=CLR_TEXT, bg=CLR_CARD,
                  activebackground=CLR_ACCENT, activeforeground=CLR_TEXT,
                  relief="flat", bd=0, padx=10, pady=4,
                  cursor="hand2",
                  command=self._show_about
                  ).pack(side="left", padx=(0, 16), pady=14)

        # Operator input
        tk.Label(top, text="Operator:", fg=CLR_MUTED, bg=CLR_PANEL,
                 font=("Arial",10)).pack(side="right", padx=(0,5))
        self.var_operator = tk.StringVar(value="Operator")
        tk.Entry(top, textvariable=self.var_operator,
                 font=("Arial",10), width=14,
                 bg=CLR_CARD, fg=CLR_TEXT, insertbackground=CLR_TEXT,
                 relief="flat", bd=4
                 ).pack(side="right", padx=(0,10), pady=12)

        # Serial number input
        tk.Label(top, text="S/N:", fg=CLR_MUTED, bg=CLR_PANEL,
                 font=("Arial",10)).pack(side="right", padx=(0,5))
        self.var_sn = tk.StringVar(value="SN-000001")
        tk.Entry(top, textvariable=self.var_sn,
                 font=("Arial",10,"bold"), width=14,
                 bg=CLR_CARD, fg=CLR_ACCENT, insertbackground=CLR_ACCENT,
                 relief="flat", bd=4
                 ).pack(side="right", padx=(0,10), pady=12)

    def _show_about(self):
        win = tk.Toplevel(self)
        win.title("About — Cable Jig Tester")
        win.resizable(False, False)
        win.configure(bg=CLR_BG)
        win.grab_set()

        # ── Header ───────────────────────────────────────────────────────
        hdr = tk.Frame(win, bg="#1A3A5C", pady=24)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚡ Cable Jig Tester",
                 font=("Arial", 20, "bold"), fg="#FFFFFF", bg="#1A3A5C"
                 ).pack()
        tk.Label(hdr, text="Cable Continuity Test System",
                 font=("Arial", 11), fg="#90B8D8", bg="#1A3A5C"
                 ).pack()

        # ── Info block ───────────────────────────────────────────────────
        info = tk.Frame(win, bg=CLR_BG, padx=32, pady=20)
        info.pack(fill="x")

        rows = [
            ("Version",      APP_TITLE),
            ("Firmware",     "dsPIC30F4011 — v4.10"),
            ("Developed by", "NEU Corporation"),
            ("Contact",      "Jasonquay@corporation.com"),
            ("Support",      "Jasonquay@corporation.com"),
            ("License",      "Proprietary — NEU Corporation"),
        ]
        for label, value in rows:
            row = tk.Frame(info, bg=CLR_BG)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=f"{label}:", width=14, anchor="e",
                     font=("Arial", 10, "bold"), fg="#90B8D8", bg=CLR_BG
                     ).pack(side="left")
            tk.Label(row, text=value, anchor="w",
                     font=("Arial", 10), fg=CLR_TEXT, bg=CLR_BG
                     ).pack(side="left", padx=(8, 0))

        # ── Divider ──────────────────────────────────────────────────────
        tk.Frame(win, bg="#2E4F6F", height=1).pack(fill="x", padx=32)

        # ── Footer ───────────────────────────────────────────────────────
        ftr = tk.Frame(win, bg=CLR_BG, pady=16)
        ftr.pack()
        import datetime
        tk.Label(ftr,
                 text=f"© {datetime.datetime.now().year} NEU Corporation. All rights reserved.",
                 font=("Arial", 9), fg=CLR_MUTED, bg=CLR_BG
                 ).pack()
        tk.Button(ftr, text="Close",
                  font=("Arial", 10), fg=CLR_TEXT, bg="#2E4F6F",
                  activebackground=CLR_ACCENT, activeforeground=CLR_TEXT,
                  relief="flat", padx=24, pady=6, bd=0, cursor="hand2",
                  command=win.destroy
                  ).pack(pady=(10, 0))

        # Centre on parent
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - win.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{x}+{y}")

    def _build_main(self):
        paned = tk.PanedWindow(self, orient="horizontal",
                               bg=CLR_BG, sashwidth=6,
                               sashrelief="flat", sashpad=2)
        paned.pack(fill="both", expand=True, padx=6, pady=4)

        left_frame = tk.Frame(paned, bg=CLR_BG)
        right_frame = tk.Frame(paned, bg=CLR_BG)
        paned.add(left_frame, minsize=300)
        paned.add(right_frame, minsize=500)

        self._build_control_panel(left_frame)
        self._build_result_panel(right_frame)

    def _card(self, parent, title):
        f = tk.LabelFrame(parent, text=f"  {title}  ",
                          bg=CLR_CARD, fg=CLR_MUTED,
                          font=("Arial",9,"bold"),
                          relief="flat", bd=2,
                          labelanchor="nw",
                          padx=8, pady=6)
        return f

    def _build_control_panel(self, parent):
        # ── Connection card ──
        conn = self._card(parent, "CONNECTION")
        conn.pack(fill="x", padx=6, pady=(6,4))

        row1 = tk.Frame(conn, bg=CLR_CARD)
        row1.pack(fill="x", pady=3)
        tk.Label(row1, text="Port:", fg=CLR_MUTED, bg=CLR_CARD,
                 font=("Arial",10), width=7, anchor="w").pack(side="left")
        self.var_port = tk.StringVar()
        self.cb_port = ttk.Combobox(row1, textvariable=self.var_port,
                                     width=16, font=("Arial",10))
        self.cb_port.pack(side="left", padx=4)

        btn_refresh = tk.Button(row1, text="⟳", command=self._refresh_ports,
                                bg=CLR_PANEL, fg=CLR_TEXT, font=("Arial",11),
                                relief="flat", cursor="hand2", padx=6)
        btn_refresh.pack(side="left")

        row2 = tk.Frame(conn, bg=CLR_CARD)
        row2.pack(fill="x", pady=3)
        self.btn_connect = tk.Button(row2, text="🔌  Connect",
                                      command=self._toggle_connect,
                                      bg=CLR_ACCENT, fg="white",
                                      font=("Arial",10,"bold"),
                                      relief="flat", cursor="hand2",
                                      padx=12, pady=6, width=16)
        self.btn_connect.pack(side="left", padx=(0,6))

        self.lbl_conn_status = tk.Label(row2, text="● Disconnected",
                                         fg=CLR_IDLE, bg=CLR_CARD,
                                         font=("Arial",10,"bold"))
        self.lbl_conn_status.pack(side="left")

        # ── Test control card ──
        ctl = self._card(parent, "TEST CONTROL")
        ctl.pack(fill="x", padx=6, pady=4)

        self.btn_test = tk.Button(ctl, text="▶  START TEST",
                                   command=self._start_test,
                                   bg="#27AE60", fg="white",
                                   font=("Arial",13,"bold"),
                                   relief="flat", cursor="hand2",
                                   padx=16, pady=12)
        self.btn_test.pack(fill="x", pady=(0,4))

        self.btn_stop = tk.Button(ctl, text="■  STOP",
                                   command=self._stop_test,
                                   bg="#7F8C8D", fg="white",
                                   font=("Arial",11,"bold"),
                                   relief="flat", cursor="hand2",
                                   padx=16, pady=8,
                                   state="disabled")
        self.btn_stop.pack(fill="x", pady=(0,6))

        self.btn_diag = tk.Button(ctl, text="🔬  Run Diagnostic (D)",
                                   command=self._run_diag,
                                   bg="#1A5276", fg="white",
                                   font=("Arial",9),
                                   relief="flat", cursor="hand2",
                                   padx=10, pady=6)
        self.btn_diag.pack(fill="x", pady=(0,2))

        self.btn_i2c_scan = tk.Button(ctl, text="🔍  I2C Bus Scan (I)",
                                       command=self._run_i2c_scan,
                                       bg="#1A5276", fg="white",
                                       font=("Arial",9),
                                       relief="flat", cursor="hand2",
                                       padx=10, pady=6)
        self.btn_i2c_scan.pack(fill="x", pady=(0,2))

        self.btn_jig_diag = tk.Button(ctl, text="🔧  Jig Self-Diag (J)",
                                       command=self._run_jig_diag,
                                       bg="#CA6F1E", fg="white",
                                       font=("Arial",9,"bold"),
                                       relief="flat", cursor="hand2",
                                       padx=10, pady=6)
        self.btn_jig_diag.pack(fill="x", pady=(0,2))

        self.btn_fdiag = tk.Button(ctl, text="🔍  F-Row Debug (F)",
                                    command=self._run_fdiag,
                                    bg="#1E8449", fg="white",
                                    font=("Arial",9),
                                    relief="flat", cursor="hand2",
                                    padx=10, pady=5)
        self.btn_fdiag.pack(fill="x", pady=(0,2))

        self.btn_j2g0 = tk.Button(ctl, text="🔎  J2 Group-0 Scan (G)",
                                   command=self._run_j2g0,
                                   bg="#1A5276", fg="white",
                                   font=("Arial",9),
                                   relief="flat", cursor="hand2",
                                   padx=10, pady=5)
        self.btn_j2g0.pack(fill="x", pady=(0,2))

        self.btn_jig_export = tk.Button(ctl, text="📋  Export Jig Report",
                                         command=self._export_jig_report,
                                         bg="#1A5276", fg="#95A5A6",
                                         font=("Arial",9),
                                         relief="flat", cursor="hand2",
                                         padx=10, pady=5,
                                         state="disabled")
        self.btn_jig_export.pack(fill="x", pady=(0,6))

        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(ctl, variable=self.progress_var,
                                         maximum=100, mode="indeterminate",
                                         length=200)
        self.progress.pack(fill="x", pady=(0,4))

        self.lbl_progress = tk.Label(ctl, text="Ready — press START",
                                      fg=CLR_MUTED, bg=CLR_CARD,
                                      font=("Arial",9), wraplength=240)
        self.lbl_progress.pack(fill="x")

        # ── Report card ──
        rpt = self._card(parent, "REPORT")
        rpt.pack(fill="x", padx=6, pady=4)

        self.btn_export = tk.Button(rpt, text="📊  Export Excel Report",
                                     command=self._export_report,
                                     bg="#8E44AD", fg="white",
                                     font=("Arial",10,"bold"),
                                     relief="flat", cursor="hand2",
                                     padx=10, pady=8,
                                     state="disabled")
        self.btn_export.pack(fill="x", pady=(0,4))

        tk.Button(rpt, text="📁  Open Reports Folder",
                  command=self._open_reports_folder,
                  bg=CLR_PANEL, fg=CLR_TEXT,
                  font=("Arial",9),
                  relief="flat", cursor="hand2",
                  padx=10, pady=6
                  ).pack(fill="x")

        # ── Stats card ──
        stat = self._card(parent, "SESSION STATS")
        stat.pack(fill="x", padx=6, pady=4)

        self.lbl_stats = tk.Label(stat,
            text="Tests: 0  |  Pass: 0  |  Fail: 0",
            fg=CLR_TEXT, bg=CLR_CARD, font=("Arial",10))
        self.lbl_stats.pack()

    def _build_result_panel(self, parent):
        # ── Big result indicator ──
        self.result_frame = tk.Frame(parent, bg=CLR_IDLE, height=80)
        self.result_frame.pack(fill="x", padx=6, pady=(6,4))
        self.result_frame.pack_propagate(False)

        self.lbl_result = tk.Label(self.result_frame,
                                    text="—  WAITING FOR TEST  —",
                                    font=("Arial",18,"bold"),
                                    fg="white", bg=CLR_IDLE)
        self.lbl_result.pack(expand=True)

        # ── Connection count badge ──
        badge_row = tk.Frame(parent, bg=CLR_BG)
        badge_row.pack(fill="x", padx=6, pady=2)

        self.badge_conn = self._badge(badge_row, "Connections Found", "0", CLR_ACCENT)
        self.badge_conn.pack(fill="x", padx=3)

        # ── Connection table ──
        tbl_frame = self._card(parent, "CONNECTIONS FOUND")
        tbl_frame.pack(fill="both", expand=True, padx=6, pady=4)

        cols = ("num", "j1", "j2")
        self.tree = ttk.Treeview(tbl_frame, columns=cols,
                                  show="headings", height=14,
                                  selectmode="browse")
        for cid, heading, w in [("num","#",50), ("j1","J1 Pin",200), ("j2","J2 Pin",200)]:
            self.tree.heading(cid, text=heading)
            self.tree.column(cid, width=w, anchor="center", minwidth=40)

        style = ttk.Style()
        style.configure("Treeview",
                         background=CLR_CARD, foreground=CLR_TEXT,
                         rowheight=24, font=("Arial",10),
                         fieldbackground=CLR_CARD)
        style.configure("Treeview.Heading",
                         background=CLR_PANEL, foreground=CLR_TEXT,
                         font=("Arial",10,"bold"))
        style.map("Treeview", background=[("selected","#2980B9")])

        sb = ttk.Scrollbar(tbl_frame, orient="vertical",
                            command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree.tag_configure("even",       background=CLR_CARD)
        self.tree.tag_configure("odd",        background="#3D5166")
        self.tree.tag_configure("gnd",        background="#7D6608", foreground="#F9E79F")
        self.tree.tag_configure("miswire",    background="#6C3483", foreground="#F8C471")
        self.tree.tag_configure("unexpected", background="#922B21", foreground="#FADBD8")

        # ── Raw log ──
        log_frame = self._card(parent, "RAW LOG")
        log_frame.pack(fill="x", padx=6, pady=(0,4))

        self.txt_log = tk.Text(log_frame, height=6, bg="#0D1117",
                                fg="#58D68D", font=("Courier",9),
                                relief="flat", wrap="word",
                                insertbackground="#58D68D")
        self.txt_log.pack(fill="both", expand=True)
        sb2 = ttk.Scrollbar(log_frame, command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb2.set)

    def _badge(self, parent, label, value, color):
        f = tk.Frame(parent, bg=color, padx=8, pady=4)
        tk.Label(f, text=value, font=("Arial",18,"bold"),
                 fg="white", bg=color).pack()
        tk.Label(f, text=label, font=("Arial",8),
                 fg="white", bg=color).pack()
        f._value_label = f.winfo_children()[0]
        return f

    def _build_statusbar(self):
        bar = tk.Frame(self, bg=CLR_PANEL, height=28)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        self.lbl_status = tk.Label(bar, text="Ready",
                                    fg=CLR_MUTED, bg=CLR_PANEL,
                                    font=("Arial",9))
        self.lbl_status.pack(side="left", padx=10, pady=4)

        self.lbl_time = tk.Label(bar, text="",
                                  fg=CLR_MUTED, bg=CLR_PANEL,
                                  font=("Arial",9))
        self.lbl_time.pack(side="right", padx=10)
        self._tick()

    def _tick(self):
        self.lbl_time.config(text=datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.after(1000, self._tick)

    # ── Actions ───────────────────────────────────────────────────────────
    def _refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.cb_port["values"] = ports
        if ports and not self.var_port.get():
            self.var_port.set(ports[0])

    def _toggle_connect(self):
        if self.serial_reader and not self.serial_stop.is_set():
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        port = self.var_port.get()
        if not port:
            messagebox.showwarning("No Port", "Please select a COM port.")
            return
        self.serial_stop.clear()
        self.serial_reader = SerialReader(port, BAUD_RATE,
                                           self.rx_queue, self.serial_stop)
        self.serial_reader.start()
        self.btn_connect.config(text="⏏  Disconnect", bg="#E74C3C")

    def _disconnect(self):
        if self.serial_reader:
            self.serial_stop.set()
            self.serial_reader = None
        self.btn_connect.config(text="🔌  Connect", bg=CLR_ACCENT)
        self.lbl_conn_status.config(text="● Disconnected", fg=CLR_IDLE)

    def _start_test(self):
        if not self.serial_reader:
            messagebox.showwarning("Not Connected",
                                   "Connect to tester first.")
            return
        if self.in_test:
            return

        sn = self.var_sn.get().strip() or "SN-000000"
        self.serial_reader.send(f"S,{sn}")
        time.sleep(0.1)
        self.serial_reader.send("T")

        self.in_test      = True
        self.test_stopped = False
        self.test_start_time = time.time()
        self.current_result = TestResult(
            serial_number = sn,
            operator      = self.var_operator.get(),
            com_port      = self.var_port.get(),
        )

        # Reset UI
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._gnd_tree_id = None
        self._gnd_count   = 0
        self._set_conn_badge(0)
        self.result_frame.config(bg="#2980B9")
        self.lbl_result.config(text="⏳  TESTING…", bg="#2980B9")
        self.btn_test.config(state="disabled")
        self.btn_stop.config(state="normal", bg="#E74C3C")
        self.btn_export.config(state="disabled")
        self.progress.config(mode="indeterminate")
        self.progress.start(15)
        self._log("=== TEST STARTED ===")

        # Start timeout watcher
        self._check_test_timeout()

    def _run_diag(self):
        """Send 'D' diagnostic — drives J1.A1 LOW, reads U19/U20 via I2C, scans all 116 J2 pins."""
        if not self.serial_reader:
            messagebox.showwarning("Not Connected", "Connect to tester first.")
            return
        if self.in_test:
            messagebox.showwarning("Test Running", "Wait for current test to finish.")
            return
        self._log("=== DIAGNOSTIC v2 SENT (D) ===")
        self._log("Checking: J1 LAT drive, U19 I2C readback, U20 I2C readback, all 116 J2 pins")
        self._log("Key results:")
        self._log("  ZA_LAT=0       → MCU driving J1 Z-line LOW correctly")
        self._log("  U19_OUT0=112   → I2C write to U19 (J1 mux) worked")
        self._log("  U20_OUT0=112   → I2C write to U20 (J2 mux) worked")
        self._log("  U19_OUT0=255   → *** U19 not responding (I2C address wrong?) ***")
        self._log("  DIAG_CONN,...  → connection detected")
        self.serial_reader.send("D")

    def _run_i2c_scan(self):
        """Send 'I' — scans I2C addresses 0x20-0x27 and reports which PCA9555 chips respond."""
        if not self.serial_reader:
            messagebox.showwarning("Not Connected", "Connect to tester first.")
            return
        if self.in_test:
            messagebox.showwarning("Test Running", "Wait for current test to finish.")
            return
        self._log("=== I2C BUS SCAN SENT (I) ===")
        self._log("Probing 0x20-0x27 — ACK = chip found, NACK = no chip")
        self._log("Expected: U19=0x20 ACK, U20=0x21 ACK (if wired correctly)")
        self.serial_reader.send("I")

    def _run_jig_diag(self):
        """Send 'J' — scan tất cả 116 J1 pin, báo cáo pin không có connection."""
        if not self.serial_reader:
            messagebox.showwarning("Not Connected", "Connect to tester first.")
            return
        if self.in_test:
            messagebox.showwarning("Test Running", "Wait for current test to finish.")
            return
        # Reset state — Pass 1 (J1)
        self._jig_fail_pins      = []
        self._jig_results_all    = []
        self._sweep_data         = {}
        self._sweep_hit_count    = 0
        self._sweep_mi           = None
        self._jig_ok_count       = 0
        self._jig_diag_ts        = datetime.now()
        self._jig_diag_com       = self.var_port.get()
        self._jig_diag_fw        = (self.current_result.fw_version
                                    if self.current_result else "")
        # Reset state — Pass 2 (J2)
        self._jig_j2_fail_pins   = []
        self._jig_j2_results_all = []
        self._sweep_j2_data      = {}
        self._sweep_j2_hit_count = 0
        self._sweep_j2_mi        = None
        self._jig_j2_ok_count    = 0
        self.btn_jig_diag.config(state="disabled")
        if hasattr(self, "btn_jig_export"):
            self.btn_jig_export.config(state="disabled", fg="#95A5A6")
        self._log("=== JIG SELF-DIAGNOSTIC SENT (J) ===")
        self._log("Pass 1: Scan 116 J1 pins (drive J1 → detect J2)")
        self._log("Pass 2: Scan 116 J2 pins (drive J2 → detect J1)")
        self._status("Jig self-diagnostic running…")
        self.serial_reader.send("J")

    def _run_j2g0(self):
        """Send 'G' — sweep J2 MUX-B group 0 (rp 0..15) from J2 side, scan J1."""
        if not self.serial_reader:
            messagebox.showwarning("Not Connected", "Connect to tester first.")
            return
        if self.in_test:
            messagebox.showwarning("Test Running", "Wait for current test to finish.")
            return
        self._j2g0_rows = []
        if hasattr(self, "btn_j2g0"):
            self.btn_j2g0.config(state="disabled")
        self._log("=== J2 GROUP-0 SWEEP SENT (G) ===")
        self._status("J2 group-0 sweep running…")
        self.serial_reader.send("G")

    def _run_fdiag(self):
        """Send 'F' — debug J1.F even pins (F2,F4,...,F16) → hiện rp thực và J2 label."""
        if not self.serial_reader:
            messagebox.showwarning("Not Connected", "Connect to tester first.")
            return
        if self.in_test:
            messagebox.showwarning("Test Running", "Wait for current test to finish.")
            return
        self._fdiag_rows = []
        if hasattr(self, "btn_fdiag"):
            self.btn_fdiag.config(state="disabled")
        self._log("=== F-ROW EVEN DEBUG SENT (F) ===")
        self._status("F-row even debug running…")
        self.serial_reader.send("F")

    def _stop_test(self):
        """User pressed STOP — cancel waiting for MCU response."""
        if not self.in_test:
            return
        self.test_stopped = True
        self._log("=== TEST STOPPED BY USER ===")
        self._status("Test stopped by user")

        # Reset UI immediately
        self.in_test = False
        self.progress.stop()
        self.progress.config(mode="determinate")
        self.progress_var.set(0)
        self.result_frame.config(bg="#7F8C8D")
        self.lbl_result.config(text="⏹  STOPPED", bg="#7F8C8D")
        self.btn_test.config(state="normal")
        self.btn_stop.config(state="disabled", bg="#7F8C8D")
        self.btn_export.config(state="disabled")

    def _check_test_timeout(self):
        """Called periodically while test is running to detect timeout."""
        if not self.in_test:
            return  # test already finished or stopped

        elapsed = time.time() - (self.test_start_time or time.time())

        # Update elapsed time in status
        self._status(f"Testing… {int(elapsed)}s elapsed  (STOP to cancel)")

        if elapsed >= self.test_timeout_s:
            # Auto-stop after timeout
            self._log(f"=== TEST TIMEOUT ({self.test_timeout_s}s) — No response from MCU ===")
            self._status(f"Timeout! No response after {self.test_timeout_s}s")
            self.test_stopped = True
            self.in_test = False
            self.progress.stop()
            self.progress.config(mode="determinate")
            self.progress_var.set(0)
            self.result_frame.config(bg=CLR_WARN)
            self.lbl_result.config(
                text=f"⚠  TIMEOUT — No response after {self.test_timeout_s}s",
                bg=CLR_WARN)
            self.btn_test.config(state="normal")
            self.btn_stop.config(state="disabled", bg="#7F8C8D")
            messagebox.showwarning("Timeout",
                f"No response from tester after {self.test_timeout_s} seconds.\n\n"
                "Possible causes:\n"
                "• I2C bus locked — power-cycle the tester\n"
                "• USB cable disconnected\n"
                "• MCU firmware crashed\n\n"
                "Check the log for the last MCU message.")
            return

        # Schedule next check in 1 second
        self.after(1000, self._check_test_timeout)

    def _set_conn_badge(self, n: int):
        self.badge_conn._value_label.config(text=str(n))

    def _log(self, line):
        self.txt_log.insert("end", line + "\n")
        self.txt_log.see("end")
        if self.current_result:
            self.current_result.raw_log.append(line)

    def _status(self, msg):
        self.lbl_status.config(text=msg)
        self.lbl_progress.config(text=msg)

    # ── Queue polling ──────────────────────────────────────────────────────
    def _poll_queue(self):
        try:
            while True:
                msg_type, data = self.rx_queue.get_nowait()
                self._handle_msg(msg_type, data)
        except queue.Empty:
            pass
        self.after(50, self._poll_queue)

    def _handle_msg(self, msg_type, data):
        if msg_type == "CONNECTED":
            self.lbl_conn_status.config(text=f"● {data}", fg=CLR_PASS)
            self._status(f"Connected: {data}")
            # Query version and ping
            if self.serial_reader:
                self.serial_reader.send("V")
                self.serial_reader.send("P")

        elif msg_type == "DISCONNECTED":
            self.lbl_conn_status.config(text="● Disconnected", fg=CLR_IDLE)
            self._status("Disconnected")

        elif msg_type == "ERROR":
            self.lbl_conn_status.config(text=f"● Error: {data}", fg=CLR_FAIL)
            self._status(f"Error: {data}")
            messagebox.showerror("Connection Error", data)

        elif msg_type == "LINE":
            self._log(data)
            self._parse_line(data)

    def _parse_line(self, line: str):
        parts = [p.strip() for p in line.split(",")]
        tag   = parts[0] if parts else ""

        if tag == "TESTER_READY":
            self._status("Tester ready")
            if self.serial_reader:
                self.serial_reader.send("V")

        elif tag == "PONG":
            self._status("Tester responded (PONG)")

        elif tag == "VERSION" and len(parts) >= 2:
            if self.current_result:
                self.current_result.fw_version = parts[1]
            self._status(f"Firmware: v{parts[1]}")

        elif tag == "PINS" and len(parts) >= 2 and parts[1].isdigit():
            if self.current_result:
                self.current_result.total_pins = int(parts[1])

        elif tag == "TEST_START" and len(parts) >= 2:
            if self.current_result:
                self.current_result.serial_number = parts[1]
            self._status(f"Scanning… S/N: {parts[1]}")

        elif tag == "CONN" and len(parts) >= 3:
            j1 = parts[1][3:] if parts[1].startswith("J1.") else parts[1]
            j2 = parts[2][3:] if parts[2].startswith("J2.") else parts[2]
            conn = ConnectionRecord(j1, j2)
            if self.current_result:
                self.current_result.connections.append(conn)
            self._add_conn_row(conn)
            self._set_conn_badge(len(self.tree.get_children()))

        elif tag == "TEST_DONE" and len(parts) >= 2:
            n = int(parts[1]) if parts[1].isdigit() else 0
            self._status(f"Scan complete — {n} connection(s) found")
            self._on_test_done()

        elif tag == "DEBUG_START":
            self._log("[DIAG] ─── Diagnostic started ───")

        elif tag == "DEBUG_DONE":
            self._log("[DIAG] ─── Diagnostic done ───")

        elif tag == "DEBUG" and len(parts) >= 2:
            if parts[1] == "CONN_FOUND":
                n = parts[2] if len(parts) >= 3 else "?"
                self._log(f"[DIAG] Connections found during diag: {n}")
                self._status(f"Diagnostic complete — {n} connection(s) found on J1 pin")
            else:
                self._log(f"[DIAG] {','.join(parts[1:])}")

        elif tag == "ZA_LAT" and len(parts) >= 2:
            if parts[1] == "0":
                self._log("[DIAG] ZA_LAT=0 → J1 Z-line LAT is LOW (za_drive_low OK)")
            else:
                self._log("[DIAG] ZA_LAT=1 → J1 Z-line LAT is HIGH (za_drive_low FAILED — MCU pin stuck?)")

        elif tag == "U19_OUT0" and len(parts) >= 2:
            v = parts[1]
            if v == "112":   # 0x70 = MUX1 enabled, channel 0
                self._log(f"[DIAG] U19_OUT0={v} (0x70) → I2C write to U19 OK, MUX1-A enabled ch=0")
            elif v == "255":  # 0xFF = unchanged = write failed
                self._log(f"[DIAG] U19_OUT0={v} (0xFF) → *** I2C WRITE TO U19 FAILED — chip not responding! ***")
            else:
                self._log(f"[DIAG] U19_OUT0={v} (unexpected value)")

        elif tag == "U20_OUT0" and len(parts) >= 2:
            v = parts[1]
            if v == "112":
                self._log(f"[DIAG] U20_OUT0={v} (0x70) → I2C write to U20 OK, MUX1-B enabled ch=0")
            elif v == "255":
                self._log(f"[DIAG] U20_OUT0={v} (0xFF) → *** I2C WRITE TO U20 FAILED — chip not responding! ***")
            else:
                self._log(f"[DIAG] U20_OUT0={v} (unexpected value)")

        elif tag == "DIAG_CONN" and len(parts) >= 3:
            self._log(f"[DIAG] CONNECTION FOUND: J1.{parts[1]} ↔ J2.{parts[2]}")

        elif tag == "I2C_SCAN_START":
            self._log("[I2C] ─── I2C bus scan started ───")

        elif tag == "I2C_SCAN_DONE":
            self._log("[I2C] ─── I2C bus scan done ───")
            self._status("I2C scan complete — check log for results")

        elif tag == "BUS1":
            self._log("[I2C] ── Bus 1 (I2C1 hardware — U19, SCL1=RF3, SDA1=RF2) ──")

        elif tag == "BUS2":
            self._log("[I2C] ── Bus 2 (I2C2 software — U20, SCL2=RD0, SDA2=RF6) ──")

        elif tag == "I2C1_PROBE" and len(parts) >= 3:
            addr_str = parts[1]
            result   = parts[2]
            try:
                addr_int = int(addr_str, 16)
            except ValueError:
                addr_int = 0
            if result == "ACK":
                note = " ← U19 (J1 mux) — OK" if addr_int == 0x20 else f" ← unexpected chip at {addr_str}"
                self._log(f"[I2C1] {addr_str} : ACK  ✓{note}")
            else:
                note = " ← U19 MISSING!" if addr_int == 0x20 else ""
                self._log(f"[I2C1] {addr_str} : NACK  —{note}")

        elif tag == "I2C2_PROBE" and len(parts) >= 3:
            addr_str = parts[1]
            result   = parts[2]
            try:
                addr_int = int(addr_str, 16)
            except ValueError:
                addr_int = 0
            if result == "ACK":
                note = " ← U20 (J2 mux) — OK" if addr_int == 0x24 else f" ← unexpected chip at {addr_str}"
                self._log(f"[I2C2] {addr_str} : ACK  ✓{note}")
            else:
                note = " ← U20 MISSING!" if addr_int == 0x24 else ""
                self._log(f"[I2C2] {addr_str} : NACK  —{note}")

        elif tag == "I2C_PROBE" and len(parts) >= 3:
            # backward compat với firmware cũ (trước khi tách I2C2)
            addr_str = parts[1]
            result   = parts[2]
            try:
                addr_int = int(addr_str, 16)
            except ValueError:
                addr_int = 0
            if result == "ACK":
                if addr_int == 0x20:   note = " ← U19 (J1 mux) — OK"
                elif addr_int == 0x24: note = " ← U20 (J2 mux) — FOUND at 0x24 (A2=1)"
                else:                  note = f" ← chip at {addr_str}"
                self._log(f"[I2C] {addr_str} : ACK  ✓{note}")
            else:
                if addr_int == 0x20: note = " ← U19 MISSING!"
                else:                note = ""
                self._log(f"[I2C] {addr_str} : NACK  —{note}")

        # ── Jig self-diagnostic (command J) ──────────────────────────────────
        elif tag == "JIG_DIAG_START":
            self._jig_fail_pins   = []
            self._jig_ok_count    = 0
            self._sweep_hit_count = 0
            self._sweep_mi        = None
            self._log("[JIG] ══════════════════════════════════════")
            self._log("[JIG] ▶  ALL-PIN SCAN  (116 J1 pins)")
            self._log("[JIG] ══════════════════════════════════════")

        elif tag == "DIAG_OK" and len(parts) >= 2:
            # DIAG_OK,<pin>,conn=<n>  — collect data, update counter silently
            pin   = parts[1]
            n_str = parts[2].replace("conn=", "") if len(parts) >= 3 else "0"
            n_val = int(n_str) if n_str.isdigit() else 0
            self._jig_results_all.append(
                JigPinResult(pin_name=pin, status="OK", conn_count=n_val))
            self._jig_ok_count += 1
            scanned = self._jig_ok_count + len(self._jig_fail_pins)
            self._status(f"Jig diagnostic: scanned {scanned}/{TOTAL_PINS}…")

        elif tag == "DIAG_FAIL" and len(parts) >= 2:
            # DIAG_FAIL,<pin>,mi=<m>,ch=<c>,LAT=<l>,OUT0=<x>,OUT1=<y>
            pin  = parts[1]
            self._jig_fail_pins.append(pin)
            info  = {k.split("=")[0]: k.split("=")[1]
                     for k in parts[2:] if "=" in k}
            mi_s   = info.get("mi",  "-1")
            ch_s   = info.get("ch",  "-1")
            lat_s  = info.get("LAT", "-1")
            out0_s = info.get("OUT0","-1")
            out1_s = info.get("OUT1","-1")
            def _toint(s):
                return int(s) if s.lstrip("-").isdigit() else -1
            mi_i, ch_i, lat_i = _toint(mi_s), _toint(ch_s), _toint(lat_s)
            out0_i, out1_i    = _toint(out0_s), _toint(out1_s)
            # Kiểm tra I2C_ERR token (firmware output ",I2C_ERR" thay OUT0/OUT1)
            has_i2c_err = any("I2C_ERR" in p for p in parts[2:])
            # Store structured result for analysis
            self._jig_results_all.append(JigPinResult(
                pin_name=pin, status="FAIL",
                mi=mi_i, ch=ch_i, lat=lat_i,
                out0=out0_i, out1=out1_i,
                i2c_err=has_i2c_err
            ))
            lat_str  = "LOW ✓" if lat_s == "0" else "HIGH ✗"
            out0_hex = f"(0x{out0_i:02X})" if out0_i >= 0 else ""
            out1_hex = f"(0x{out1_i:02X})" if out1_i >= 0 else ""
            self._log(f"[JIG] ✗  {pin:<4}  mi={mi_s} ch={ch_s}"
                      f"  LAT={lat_s}={lat_str}"
                      f"  OUT0={out0_s}{out0_hex}"
                      f"  OUT1={out1_s}{out1_hex}")
            if lat_s == "1":
                self._log(f"[JIG]     ⚠ Z-line drive FAIL (mi={mi_s}) — MCU port stuck?")
            if out0_s == "255" or out1_s == "255":
                self._log(f"[JIG]     ⚠ OUT=255 → I2C write FAILED (U19 not responding)")

        elif tag == "SWEEP_START" and len(parts) >= 2:
            mi_val = parts[1].replace("mi=", "")
            self._sweep_mi        = mi_val
            self._sweep_hit_count = 0
            # Initialize entry so empty sweep ≠ "never swept"
            self._sweep_data.setdefault(mi_val, set())
            self._log(f"[JIG]     ─── Sweep MUX group {mi_val} (16 ch) ───────────────")

        elif tag == "SWEEP_HIT" and len(parts) >= 3:
            self._sweep_hit_count += 1
            ch_val = parts[1].replace("ch=", "")
            j2_val = parts[2].replace("J2=", "")
            # Track which channels produced a hit in this group
            if self._sweep_mi is not None and ch_val.isdigit():
                self._sweep_data.setdefault(self._sweep_mi, set()).add(int(ch_val))
            self._log(f"[JIG]         ch={ch_val:>2} → J2.{j2_val}")

        elif tag == "SWEEP_END" and len(parts) >= 2:
            mi_val = parts[1].replace("mi=", "")
            n_hit  = self._sweep_hit_count
            if n_hit == 0:
                self._log(f"[JIG]     ─── group {mi_val}: 0 hits → IC chết hoặc Z-line đứt")
            else:
                self._log(f"[JIG]     ─── group {mi_val}: {n_hit} ch OK, còn lại dead → trace đứt")
            self._sweep_mi        = None
            self._sweep_hit_count = 0

        # ── J2 pass tags ─────────────────────────────────────────────────────
        elif tag == "JIG_DIAG_J2_START":
            self._jig_j2_fail_pins   = []
            self._jig_j2_ok_count    = 0
            self._sweep_j2_hit_count = 0
            self._sweep_j2_mi        = None
            self._log("[JIG] ══════════════════════════════════════")
            self._log("[JIG] ▶  PASS 2 — J2 SIDE SCAN  (116 J2 pins)")
            self._log("[JIG] ══════════════════════════════════════")

        elif tag == "DIAG_J2_OK" and len(parts) >= 2:
            pin   = parts[1]
            n_str = parts[2].replace("conn=", "") if len(parts) >= 3 else "0"
            n_val = int(n_str) if n_str.isdigit() else 0
            self._jig_j2_results_all.append(
                JigPinResult(pin_name=pin, status="OK", conn_count=n_val))
            self._jig_j2_ok_count += 1
            scanned = self._jig_j2_ok_count + len(self._jig_j2_fail_pins)
            self._status(f"Jig diagnostic P2: scanned {scanned}/{TOTAL_PINS}…")

        elif tag == "DIAG_J2_FAIL" and len(parts) >= 2:
            # DIAG_J2_FAIL,<pin>,mi=<m>,ch=<c>,LAT=<l>,OUT0=<x>,OUT1=<y>
            pin  = parts[1]
            self._jig_j2_fail_pins.append(pin)
            info  = {k.split("=")[0]: k.split("=")[1]
                     for k in parts[2:] if "=" in k}
            mi_s   = info.get("mi",  "-1")
            ch_s   = info.get("ch",  "-1")
            lat_s  = info.get("LAT", "-1")
            out0_s = info.get("OUT0","-1")
            out1_s = info.get("OUT1","-1")
            def _toint2(s):
                return int(s) if s.lstrip("-").isdigit() else -1
            mi_i, ch_i, lat_i = _toint2(mi_s), _toint2(ch_s), _toint2(lat_s)
            out0_i, out1_i    = _toint2(out0_s), _toint2(out1_s)
            has_i2c_err = any("I2C_ERR" in p for p in parts[2:])
            self._jig_j2_results_all.append(JigPinResult(
                pin_name=pin, status="FAIL",
                mi=mi_i, ch=ch_i, lat=lat_i,
                out0=out0_i, out1=out1_i,
                i2c_err=has_i2c_err
            ))
            lat_str  = "LOW ✓" if lat_s == "0" else "HIGH ✗"
            out0_hex = f"(0x{out0_i:02X})" if out0_i >= 0 else ""
            out1_hex = f"(0x{out1_i:02X})" if out1_i >= 0 else ""
            self._log(f"[J2]  ✗  {pin:<4}  mi={mi_s} ch={ch_s}"
                      f"  LAT={lat_s}={lat_str}"
                      f"  OUT0={out0_s}{out0_hex}"
                      f"  OUT1={out1_s}{out1_hex}")

        elif tag == "SWEEP_J2_START" and len(parts) >= 2:
            mi_val = parts[1].replace("mi=", "")
            self._sweep_j2_mi        = mi_val
            self._sweep_j2_hit_count = 0
            self._sweep_j2_data.setdefault(mi_val, set())
            self._log(f"[J2]      ─── Sweep J2 MUX group {mi_val} (16 ch) ───────────")

        elif tag == "SWEEP_J2_HIT" and len(parts) >= 3:
            self._sweep_j2_hit_count += 1
            ch_val = parts[1].replace("ch=", "")
            j1_val = parts[2].replace("J1=", "")
            if self._sweep_j2_mi is not None and ch_val.isdigit():
                self._sweep_j2_data.setdefault(self._sweep_j2_mi, set()).add(int(ch_val))
            self._log(f"[J2]          ch={ch_val:>2} → J1.{j1_val}")

        elif tag == "SWEEP_J2_END" and len(parts) >= 2:
            mi_val = parts[1].replace("mi=", "")
            n_hit  = self._sweep_j2_hit_count
            if n_hit == 0:
                self._log(f"[J2]      ─── group {mi_val}: 0 hits → IC chết hoặc Z-line đứt")
            else:
                self._log(f"[J2]      ─── group {mi_val}: {n_hit} ch OK, còn lại dead → trace đứt")
            self._sweep_j2_mi        = None
            self._sweep_j2_hit_count = 0

        elif tag == "JIG_DIAG_DONE":
            # Parse j1=N,j2=M  (or legacy fail=N for backward compat)
            j1_n = 0;  j2_n = 0
            for p in parts[1:]:
                if p.startswith("j1="):
                    try: j1_n = int(p[3:])
                    except ValueError: pass
                elif p.startswith("j2="):
                    try: j2_n = int(p[3:])
                    except ValueError: pass
                elif p.startswith("fail="):       # legacy firmware
                    try: j1_n = int(p[5:])
                    except ValueError: pass
            fail_n = j1_n + j2_n

            # ── Root cause analysis ────────────────────────────────────────
            self._analyze_jig_results()
            self._analyze_jig_j2_results()

            # ── Log summary ────────────────────────────────────────────────
            self._log("[JIG] ══════════════════════════════════════")
            if j1_n == 0:
                self._log("[JIG] ✓  J1 ALL CLEAR — 116/116 pins OK")
            else:
                self._log(f"[JIG] ✗  J1: {j1_n} pin(s) FAILED — ROOT CAUSE:")
                for r in self._jig_results_all:
                    if r.status == "FAIL":
                        self._log(f"[JIG]   • J1.{r.pin_name:<4}: {r.root_cause}")
                        self._log(f"[JIG]          → {r.recommendation}")
            if j2_n == 0:
                self._log("[JIG] ✓  J2 ALL CLEAR — 116/116 pins OK")
            else:
                self._log(f"[JIG] ✗  J2: {j2_n} pin(s) FAILED — ROOT CAUSE:")
                for r in self._jig_j2_results_all:
                    if r.status == "FAIL":
                        self._log(f"[JIG]   • J2.{r.pin_name:<4}: {r.root_cause}")
                        self._log(f"[JIG]          → {r.recommendation}")
            self._log("[JIG] ══════════════════════════════════════")
            if fail_n == 0:
                self._status("Jig diagnostic: J1+J2 tất cả OK ✓")
            else:
                self._status(f"Jig diagnostic: J1={j1_n} fail, J2={j2_n} fail")

            # ── Enable buttons ─────────────────────────────────────────────
            if hasattr(self, "btn_jig_diag"):
                self.btn_jig_diag.config(state="normal")
            if hasattr(self, "btn_jig_export"):
                self.btn_jig_export.config(state="normal", fg="white")

            # ── Popup ──────────────────────────────────────────────────────
            if fail_n == 0:
                messagebox.showinfo("Jig Self-Diagnostic",
                                    "✓  Tất cả 116 pin J1 + 116 pin J2 đều OK!\n"
                                    "Jig hardware hoàn toàn bình thường.")
            else:
                rows_j1 = [r for r in self._jig_results_all       if r.status == "FAIL"]
                rows_j2 = [r for r in self._jig_j2_results_all    if r.status == "FAIL"]
                detail_j1 = "\n".join(
                    f"  ✗ J1.{r.pin_name}: {r.root_cause}" for r in rows_j1)
                detail_j2 = "\n".join(
                    f"  ✗ J2.{r.pin_name}: {r.root_cause}" for r in rows_j2)
                body = ""
                if rows_j1:
                    body += f"J1 ({j1_n} pin):\n{detail_j1}\n\n"
                if rows_j2:
                    body += f"J2 ({j2_n} pin):\n{detail_j2}\n\n"
                messagebox.showwarning(
                    "Jig Self-Diagnostic — Phân tích nguyên nhân",
                    f"Tổng cộng {fail_n} pin thất bại:\n\n{body}"
                    "Nhấn 'Export Jig Report' để lưu báo cáo Excel đầy đủ."
                )

        elif tag == "J2G0_START":
            self._j2g0_rows = []
            self._log("[J2G0] ══ J2 MUX-B group-0 sweep (rp 0..15) ══")

        elif tag == "J2G0":
            # J2G0,rp=<n>,J1.<pin>  or  J2G0,rp=<n>,NO_CONN
            rp_str = parts[1].replace("rp=", "") if len(parts) >= 2 else "?"
            j1_pin = parts[2].replace("J1.", "") if len(parts) >= 3 else "NO_CONN"
            if j1_pin == "NO_CONN":
                self._log(f"[J2G0] rp={rp_str:>2} → NO_CONN")
            else:
                self._log(f"[J2G0] rp={rp_str:>2} → J1.{j1_pin}")
            self._j2g0_rows.append((rp_str, j1_pin))

        elif tag == "J2G0_END":
            connected = [(rp, j1) for rp, j1 in self._j2g0_rows if j1 != "NO_CONN"]
            self._log(f"[J2G0] {len(connected)}/16 channels có kết nối J1")
            self._log("[J2G0] ══════════════════════════════")
            if hasattr(self, "btn_j2g0"):
                self.btn_j2g0.config(state="normal")

        elif tag == "FDIAG_START":
            self._fdiag_rows = []
            self._log("[FDIAG] ══ F-row even-pin debug ══")

        elif tag == "FDIAG":
            # FDIAG,J1.F2,rp=33,J2.F2   or   FDIAG,J1.F2,NO_CONN
            if len(parts) >= 3:
                j1_pin = parts[1].replace("J1.", "")
                if parts[2] == "NO_CONN":
                    self._log(f"[FDIAG] J1.{j1_pin:>3} → NO_CONN  ✗")
                    self._fdiag_rows.append((j1_pin, None, None, "FAIL"))
                else:
                    rp_str  = parts[2].replace("rp=", "")
                    j2_pin  = parts[3].replace("J2.", "") if len(parts) >= 4 else "?"
                    self._log(f"[FDIAG] J1.{j1_pin:>3} → J2.{j2_pin:<4}  (MUX-B rp={rp_str})")
                    self._fdiag_rows.append((j1_pin, j2_pin, rp_str, "OK"))

        elif tag == "FDIAG_END":
            ok  = sum(1 for r in self._fdiag_rows if r[3] == "OK")
            nok = sum(1 for r in self._fdiag_rows if r[3] == "FAIL")
            self._log(f"[FDIAG] Kết quả: {ok} OK, {nok} NO_CONN")
            self._log("[FDIAG] ══════════════════════════════")
            if nok:
                fails = ", ".join(f"J1.F{r[0]}" for r in self._fdiag_rows if r[3]=="FAIL")
                messagebox.showwarning("F-Row Debug",
                    f"{nok} pin không tìm thấy kết nối J2:\n{fails}\n\n"
                    "Kiểm tra PCB J2 MUX-B nhóm 0 (rp=0..15).")
            if hasattr(self, "btn_fdiag"):
                self.btn_fdiag.config(state="normal")

        elif tag == "JIG_DIAG_I2C_ERR":
            drv = parts[1].replace("drv=", "") if len(parts) >= 2 else "?"
            self._log(f"[JIG] ⚠  I2C error J1 (drv={drv}) — pin bị bỏ qua")

        elif tag == "JIG_DIAG_J2_I2C_ERR":
            drv = parts[1].replace("drv=", "") if len(parts) >= 2 else "?"
            self._log(f"[JIG] ⚠  I2C error J2 (drv={drv}) — pin bị bỏ qua")

        elif tag == "SWEEP_I2C_ERR":
            ch = parts[1].replace("ch=", "") if len(parts) >= 2 else "?"
            self._log(f"[JIG] ⚠  I2C error trong sweep J1 (ch={ch}) — dừng sweep")
            if hasattr(self, "btn_jig_diag"):
                self.btn_jig_diag.config(state="normal")

        elif tag == "SWEEP_J2_I2C_ERR":
            ch = parts[1].replace("ch=", "") if len(parts) >= 2 else "?"
            self._log(f"[JIG] ⚠  I2C error trong sweep J2 (ch={ch}) — dừng sweep")
            if hasattr(self, "btn_jig_diag"):
                self.btn_jig_diag.config(state="normal")

        elif tag == "ERROR":
            err_code = parts[1] if len(parts) >= 2 else "UNKNOWN"
            self._log(f"=== MCU ERROR: {err_code} ===")
            if self.in_test:
                self.test_stopped = True
                self.in_test = False
                self.progress.stop()
                self.progress.config(mode="determinate")
                self.progress_var.set(0)
                self.result_frame.config(bg=CLR_FAIL)
                self.lbl_result.config(text=f"⚠  ERROR: {err_code}", bg=CLR_FAIL)
                self.btn_test.config(state="normal")
                self.btn_stop.config(state="disabled", bg="#7F8C8D")
                msg = {
                    "I2C_TIMEOUT":   "I2C bus is locked (SDA stuck low).\n"
                                     "Power-cycle the tester and check the\n"
                                     "PCA9555 expanders (U19 on I2C1, U20 on I2C2).",
                    "I2C_INIT_FAIL": "I2C bus failed during startup.\n"
                                     "Check PCA9555 power and wiring:\n"
                                     "U19: I2C1 (RF3=SCL, RF2=SDA)\n"
                                     "U20: I2C2 soft (RD0=SCL2, RF6=SDA2).",
                }.get(err_code, f"MCU reported error: {err_code}")
                messagebox.showerror("MCU Hardware Error", msg)
            else:
                self._status(f"MCU error: {err_code}")

    def _add_conn_row(self, conn: ConnectionRecord):
        j1, j2 = conn.j1_pin, conn.j2_pin

        # GND net — collapse to one summary row
        if j1 in GND_J1 or j2 in GND_J2:
            self._gnd_count += 1
            if self._gnd_tree_id is None:
                idx = len(self.tree.get_children()) + 1
                self._gnd_tree_id = self.tree.insert(
                    "", "end",
                    values=(idx, "GND bus", f"GND — {self._gnd_count} kết nối"),
                    tags=("gnd",))
            else:
                vals = self.tree.item(self._gnd_tree_id, "values")
                self.tree.item(self._gnd_tree_id,
                               values=(vals[0], "GND bus",
                                       f"GND — {self._gnd_count} kết nối"))
            return

        # Signal connection — validate
        idx = len(self.tree.get_children()) + 1
        expected_j2 = EXPECTED_J1_TO_J2.get(j1)
        if expected_j2 is None:
            tag_name = "unexpected"
            j2_disp  = f"J2.{j2}  ⚠ unexpected"
        elif j2 == expected_j2:
            tag_name = "odd" if idx % 2 else "even"
            j2_disp  = f"J2.{j2}  ✓"
        else:
            tag_name = "miswire"
            j2_disp  = f"J2.{j2}  ✗ expected J2.{expected_j2}"
        self.tree.insert("", "end",
                          values=(idx, f"J1.{j1}", j2_disp),
                          tags=(tag_name,))

    def _on_test_done(self):
        if self.test_stopped:
            return

        self.in_test = False
        self.progress.stop()
        self.progress.config(mode="determinate")
        self.progress_var.set(100)

        n = 0
        missing_signals: list = []
        miswire_signals: list = []

        if self.current_result:
            self.current_result.completed = True
            if self.test_start_time:
                self.current_result.duration_s = time.time() - self.test_start_time
            n = len(self.current_result.connections)
            self.history.append(self.current_result)
            self._save_history()
            self._update_stats()

            # ── Build signal detection with SHORT fault detection ─────────
            # For each signal J1 pin:
            #   - Collect its non-GND J2 hits (the cable signal connections)
            #   - Check if it also has GND_J2 hits (= shorted to GND bus)
            # Use same skip-GND logic as generate() for signal classification,
            # but additionally flag pins that appear in BOTH GND and signal hits.

            # Group connections by J1 pin
            from collections import defaultdict
            j1_gnd_hits  = defaultdict(list)  # signal J1 pins that also hit GND J2
            j1_sig_hits  = defaultdict(list)  # signal J1 pins → their signal J2 hits

            for c in self.current_result.connections:
                if c.j1_pin in GND_J1:
                    continue
                if c.j2_pin in GND_J2:
                    j1_gnd_hits[c.j1_pin].append(c.j2_pin)  # GND hit on a signal J1 pin
                else:
                    j1_sig_hits[c.j1_pin].append(c.j2_pin)  # real signal hit

            # Build detected: first signal hit per J1 pin
            detected: dict = {}
            for j1, hits in j1_sig_hits.items():
                if hits:
                    detected[j1] = hits[0]

            # Detect SHORT faults: signal J1 pin also pulled GND bus
            short_signals: list = []
            for j1 in sorted(j1_gnd_hits.keys(), key=lambda x:(x[0],int(x[1:]))):
                if j1_gnd_hits[j1]:  # has GND hits = shorted to GND bus
                    # Find which GND J1 pin it's shorted to (same row, adjacent pin)
                    short_signals.append(f"J1.{j1} shorted to GND bus")
            for j1, exp_j2 in EXPECTED_J1_TO_J2.items():
                got = detected.get(j1)
                if got is None:
                    missing_signals.append(f"J1.{j1}→J2.{exp_j2}")
                elif got != exp_j2:
                    miswire_signals.append(f"J1.{j1}: got J2.{got} exp J2.{exp_j2}")

        short_count  = len(short_signals)
        pass_ok = not missing_signals and not miswire_signals and not short_signals
        if pass_ok:
            self.result_frame.config(bg=CLR_PASS)
            self.lbl_result.config(text=f"✓  PASS — {n} connection(s) verified", bg=CLR_PASS)
            self._status(f"PASS — {n} connections OK — press Export to save")
        else:
            issues = len(missing_signals) + len(miswire_signals) + short_count
            self.result_frame.config(bg=CLR_FAIL)
            self.lbl_result.config(text=f"✗  FAIL — {issues} issue(s) — check log", bg=CLR_FAIL)
            self._status(f"FAIL — {issues} issue(s) detected")
            if short_signals:
                self._log(f"[RESULT] SHORT ({short_count}): "
                          + ", ".join(short_signals[:10])
                          + (" …" if short_count > 10 else ""))
            if missing_signals:
                self._log(f"[RESULT] MISSING ({len(missing_signals)}): "
                          + ", ".join(missing_signals[:10])
                          + (" …" if len(missing_signals) > 10 else ""))
            if miswire_signals:
                self._log(f"[RESULT] MISWIRE ({len(miswire_signals)}): "
                          + ", ".join(miswire_signals[:10])
                          + (" …" if len(miswire_signals) > 10 else ""))

        self.btn_test.config(state="normal")
        self.btn_stop.config(state="disabled", bg="#7F8C8D")
        self.btn_export.config(state="normal" if EXCEL_OK else "disabled")

        sn = self.var_sn.get()
        try:
            prefix = sn.rstrip("0123456789")
            num    = int(sn[len(prefix):]) + 1
            self.var_sn.set(f"{prefix}{num:06d}")
        except Exception:
            pass

    def _update_stats(self):
        total = len(self.history)
        done  = sum(1 for r in self.history if r.completed)
        self.lbl_stats.config(
            text=f"Scans: {total}  |  Completed: {done}")

    # ── Jig Diagnostic Analysis & Export ─────────────────────────────────────
    def _analyze_jig_results(self):
        """Phân tích nguyên nhân cho từng pin FAIL dựa trên sweep data.

        Thứ tự ưu tiên:
          1. i2c_err  → firmware báo I2C lỗi khi đọc thanh ghi
          2. lat==1   → MCU không drive được Z-line xuống LOW
          3. OUT0/OUT1 != expected → I2C write không được ghi nhận
          4. sweep data → trace đứt / IC chết / pogo xấu

        LƯU Ý QUAN TRỌNG:
          OUT1=0xFF (255) là GIÁ TRỊ ĐÚNG cho IO0-based group (mi=0,1,4,5)
          vì mux_a() luôn set u19_io1=0xFF cho những group này.
          KHÔNG được dùng OUT1==255 làm chỉ báo I2C lỗi.
        """
        EN_IO0 = {0: 0x80, 1: 0x20, 4: 0x40, 5: 0x10}  # IO0-based groups
        EN_IO1 = {2: 0x80, 3: 0x20, 6: 0x40, 7: 0x10}  # IO1-based groups

        def expected_regs(mi, ch):
            """Giá trị OUT0/OUT1 mong đợi sau mux_a(mi, ch) thành công."""
            ch = ch & 0x0F
            if mi in EN_IO0:
                return (0xF0 | ch) & (~EN_IO0[mi] & 0xFF), 0xFF
            if mi in EN_IO1:
                return 0xF0 | ch, (~EN_IO1[mi]) & 0xFF
            return -1, -1

        for r in self._jig_results_all:
            if r.status == "OK":
                r.root_cause     = "Bình thường"
                r.recommendation = "—"
                continue

            # ── 1. I2C error được firmware báo tường minh ──────────────────
            if r.i2c_err:
                r.root_cause = "I2C error khi đọc lại thanh ghi U19"
                r.recommendation = ("Kiểm tra PCA9555 U19: nguồn VCC, "
                                    "địa chỉ I2C (0x20), dây SDA/SCL (RF2/RF3)")

            # ── 2. MCU không drive được Z-line ─────────────────────────────
            elif r.lat == 1:
                r.root_cause = "MCU Z-line drive thất bại (LAT=HIGH)"
                r.recommendation = (f"Kiểm tra MCU port pin (group mi={r.mi}): "
                                    "dsPIC RE/RF bị ngắn mạch hoặc lỗi firmware")

            # ── 3. Không có dữ liệu đủ để phân tích ───────────────────────
            elif r.mi < 0 or r.out0 < 0:
                r.root_cause     = "Không đủ dữ liệu để phân tích"
                r.recommendation = "Chạy lại Jig Diagnostic"

            else:
                # ── 4. So sánh OUT0/OUT1 thực tế với giá trị mong đợi ──────
                exp0, exp1 = expected_regs(r.mi, r.ch)
                regs_match = (exp0 < 0) or (r.out0 == exp0 and r.out1 == exp1)

                if not regs_match:
                    r.root_cause = (
                        f"Thanh ghi MUX sai sau mux_a({r.mi},{r.ch}): "
                        f"OUT0={r.out0} (cần {exp0}), "
                        f"OUT1={r.out1} (cần {exp1})"
                    )
                    r.recommendation = ("Kiểm tra U19 I2C: nguồn VCC, địa chỉ I2C; "
                                        "thử tăng I2C clock delay trong firmware")

                else:
                    # ── 5. Drive + MUX config đúng → hardware jig ──────────
                    mi_key = str(r.mi)
                    sweep  = self._sweep_data.get(mi_key, None)

                    if sweep is None:
                        r.root_cause     = ("Không có continuity "
                                            "— chưa có dữ liệu sweep")
                        r.recommendation = "Chạy lại Jig Diagnostic"

                    elif len(sweep) == 0:
                        r.root_cause = (
                            f"Toàn bộ MUX group {r.mi} không hoạt động "
                            f"(0/16 channel có continuity)"
                        )
                        r.recommendation = (
                            f"Kiểm tra IC MUX group {r.mi}: "
                            f"(1) Nguồn VCC cấp IC, "
                            f"(2) Chân EN (active LOW từ U19), "
                            f"(3) Z-line trên PCB"
                        )

                    elif r.ch not in sweep:
                        working = sorted(sweep)
                        r.root_cause = (
                            f"PCB trace đứt: MUX group {r.mi} "
                            f"output Y{r.ch} → pogo pin J1.{r.pin_name}"
                        )
                        r.recommendation = (
                            f"Vá/kiểm tra trace PCB từ MUX group {r.mi} "
                            f"output Y{r.ch} đến pogo J1.{r.pin_name}  "
                            f"(ch {working} trong cùng group vẫn OK)"
                        )

                    else:
                        r.root_cause = (
                            f"Pogo pin J1.{r.pin_name} tiếp xúc kém "
                            f"(ch={r.ch} có tín hiệu trong sweep)"
                        )
                        r.recommendation = (
                            f"Kiểm tra/thay pogo pin J1.{r.pin_name}; "
                            f"hoặc tăng settle delay "
                            f"(delay_us 200 → 500)"
                        )

    def _analyze_jig_j2_results(self):
        """Phân tích nguyên nhân cho từng pin J2 FAIL (PASS 2).

        Logic giống _analyze_jig_results() nhưng:
          • Dữ liệu: _jig_j2_results_all + _sweep_j2_data
          • Thanh ghi: U20 (địa chỉ 0x24, cấu trúc giống U19)
          • Z-line: PORTB (thay PORTE/PORTF)
          • OUT1=0xFF là đúng cho IO0-based group (mi=0,1,4,5) — quy tắc giống U19

        LƯU Ý: expected_regs() dùng cùng công thức vì U20 có PCA9555 giống hệt U19.
        """
        EN_IO0 = {0: 0x80, 1: 0x20, 4: 0x40, 5: 0x10}
        EN_IO1 = {2: 0x80, 3: 0x20, 6: 0x40, 7: 0x10}

        def expected_regs(mi, ch):
            ch = ch & 0x0F
            if mi in EN_IO0:
                return (0xF0 | ch) & (~EN_IO0[mi] & 0xFF), 0xFF
            if mi in EN_IO1:
                return 0xF0 | ch, (~EN_IO1[mi]) & 0xFF
            return -1, -1

        for r in self._jig_j2_results_all:
            if r.status == "OK":
                r.root_cause     = "Bình thường"
                r.recommendation = "—"
                continue

            # ── 1. I2C error ────────────────────────────────────────────────
            if r.i2c_err:
                r.root_cause = "I2C2 error khi đọc lại thanh ghi U20"
                r.recommendation = ("Kiểm tra PCA9555 U20: nguồn VCC, "
                                    "địa chỉ I2C (0x24), dây SDA2/SCL2 (RF6/RD0)")

            # ── 2. MCU không drive được Z-line ─────────────────────────────
            elif r.lat == 1:
                r.root_cause = "MCU Z-line drive thất bại (LAT=HIGH)"
                r.recommendation = (f"Kiểm tra MCU PORTB pin (group mi={r.mi}): "
                                    "dsPIC RB bị ngắn mạch hoặc lỗi firmware")

            # ── 3. Không đủ dữ liệu ─────────────────────────────────────────
            elif r.mi < 0 or r.out0 < 0:
                r.root_cause     = "Không đủ dữ liệu để phân tích"
                r.recommendation = "Chạy lại Jig Diagnostic"

            else:
                # ── 4. So sánh OUT0/OUT1 thực tế với giá trị mong đợi ──────
                exp0, exp1 = expected_regs(r.mi, r.ch)
                regs_match = (exp0 < 0) or (r.out0 == exp0 and r.out1 == exp1)

                if not regs_match:
                    r.root_cause = (
                        f"Thanh ghi MUX sai sau mux_b({r.mi},{r.ch}): "
                        f"OUT0={r.out0} (cần {exp0}), "
                        f"OUT1={r.out1} (cần {exp1})"
                    )
                    r.recommendation = ("Kiểm tra U20 I2C2: nguồn VCC, địa chỉ I2C (0x24); "
                                        "thử tăng delay trong software I2C2 (RD0/RF6)")

                else:
                    # ── 5. Drive + MUX config đúng → phân tích sweep ───────
                    mi_key = str(r.mi)
                    sweep  = self._sweep_j2_data.get(mi_key, None)

                    if sweep is None:
                        r.root_cause     = ("Không có continuity "
                                            "— chưa có dữ liệu sweep J2")
                        r.recommendation = "Chạy lại Jig Diagnostic"

                    elif len(sweep) == 0:
                        r.root_cause = (
                            f"Toàn bộ J2 MUX group {r.mi} không hoạt động "
                            f"(0/16 channel có continuity)"
                        )
                        r.recommendation = (
                            f"Kiểm tra IC MUX B-side group {r.mi}: "
                            f"(1) Nguồn VCC, "
                            f"(2) Chân EN (active LOW từ U20), "
                            f"(3) Z-line PORTB trên PCB"
                        )

                    elif r.ch not in sweep:
                        working = sorted(sweep)
                        r.root_cause = (
                            f"PCB trace đứt: J2 MUX group {r.mi} "
                            f"output Y{r.ch} → pogo pin J2.{r.pin_name}"
                        )
                        r.recommendation = (
                            f"Vá/kiểm tra trace PCB từ J2 MUX group {r.mi} "
                            f"output Y{r.ch} đến pogo J2.{r.pin_name}  "
                            f"(ch {working} trong cùng group vẫn OK)"
                        )

                    else:
                        r.root_cause = (
                            f"Pogo pin J2.{r.pin_name} tiếp xúc kém "
                            f"(ch={r.ch} có tín hiệu trong sweep)"
                        )
                        r.recommendation = (
                            f"Kiểm tra/thay pogo pin J2.{r.pin_name}; "
                            f"hoặc tăng settle delay "
                            f"(delay_us 200 → 500)"
                        )

    def _export_jig_report(self):
        """Xuất kết quả Jig Self-Diagnostic ra file Excel (J1 + J2)."""
        if not EXCEL_OK:
            messagebox.showerror("Missing Library",
                                 "openpyxl chưa cài.\npip install openpyxl")
            return
        if not self._jig_results_all and not self._jig_j2_results_all:
            messagebox.showinfo("No Data", "Chạy Jig Self-Diagnostic trước.")
            return

        wb = openpyxl.Workbook()

        # Sheet 1 — J1: tất cả 116 pin
        ws1 = wb.active
        ws1.title = "J1 All Pins"
        if self._jig_results_all:
            self._build_jig_sheet_all(ws1, self._jig_results_all, side_label="J1")
        else:
            ws1["A1"] = "Không có dữ liệu J1"

        # Sheet 2 — J1: chỉ pin FAIL (nếu có)
        j1_failed = [r for r in self._jig_results_all if r.status == "FAIL"]
        if j1_failed:
            ws2 = wb.create_sheet("J1 Root Cause")
            self._build_jig_sheet_fail(ws2, j1_failed, side_label="J1")

        # Sheet 3 — J2: tất cả 116 pin
        ws3 = wb.create_sheet("J2 All Pins")
        if self._jig_j2_results_all:
            self._build_jig_sheet_all(ws3, self._jig_j2_results_all, side_label="J2")
        else:
            ws3["A1"] = "Không có dữ liệu J2"

        # Sheet 4 — J2: chỉ pin FAIL (nếu có)
        j2_failed = [r for r in self._jig_j2_results_all if r.status == "FAIL"]
        if j2_failed:
            ws4 = wb.create_sheet("J2 Root Cause")
            self._build_jig_sheet_fail(ws4, j2_failed, side_label="J2")

        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        ts    = (self._jig_diag_ts or datetime.now()).strftime("%Y%m%d_%H%M%S")
        fname = f"jig_diag_{ts}.xlsx"
        out   = reports_dir / fname

        try:
            wb.save(str(out))
        except Exception as e:
            messagebox.showerror("Export Error", str(e))
            return

        self._log(f"[JIG] 📋  Report saved: reports/{fname}")
        if messagebox.askyesno("Đã lưu", f"Lưu thành công:\n{fname}\n\nMở ngay?"):
            if sys.platform == "win32":
                os.startfile(str(out))
            else:
                os.system(f"xdg-open '{out}'")

    # ── Excel sheet builders ──────────────────────────────────────────────
    def _jig_thin(self):
        s = Side(style='thin', color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def _jig_fill(self, c):
        return PatternFill("solid", fgColor=c)

    def _build_jig_sheet_all(self, ws, results, side_label="J1"):
        """Sheet 'All Pins' — toàn bộ 116 pin với status + drive state + analysis."""
        thin = self._jig_thin
        fill = self._jig_fill
        ts   = (self._jig_diag_ts or datetime.now()).strftime("%Y-%m-%d  %H:%M:%S")
        ok_n  = sum(1 for r in results if r.status == "OK")
        fail_n = len(results) - ok_n

        ws.sheet_view.showGridLines = False

        # ── Title ──────────────────────────────────────────────────────────
        ws.merge_cells("A1:K1")
        c = ws["A1"]
        c.value     = f"JIG SELF-DIAGNOSTIC REPORT — {side_label} SIDE"
        c.font      = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        c.fill      = fill("1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # ── Info ────────────────────────────────────────────────────────────
        for lbl_cell, lbl_val, val_cell, val_val in [
            ("A2", "Date / Time:", "B2", ts),
            ("D2", "COM Port:",    "E2", self._jig_diag_com or "—"),
            ("G2", "Firmware:",    "H2", self._jig_diag_fw  or "—"),
        ]:
            ws[lbl_cell].value = lbl_val
            ws[lbl_cell].font  = Font(name="Arial", size=10, bold=True)
            ws[lbl_cell].alignment = Alignment(horizontal="right")
            ws[val_cell].value = val_val
            ws[val_cell].font  = Font(name="Arial", size=10)
        ws.row_dimensions[2].height = 18

        # ── Summary badges ──────────────────────────────────────────────────
        for ref, txt, clr in [("A3", f"Total: {len(results)}", "1A5276"),
                               ("C3", f"OK: {ok_n}",           "1E8449"),
                               ("E3", f"FAIL: {fail_n}",       "922B21")]:
            ws.merge_cells(f"{ref}:{chr(ord(ref[0])+1)}3")
            c = ws[ref]
            c.value     = txt
            c.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            c.fill      = fill(clr)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22

        # ── Column headers ──────────────────────────────────────────────────
        HDR    = ["#","Pin","Status","MI","CH","Conn","LAT","OUT0","OUT1",
                  "Root Cause","Recommendation"]
        WIDTHS = [ 5,   8,     8,     6,   6,   6,    6,    8,    8,     38,   46]
        for col, (h, w) in enumerate(zip(HDR, WIDTHS), 1):
            c = ws.cell(5, col, h)
            c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill      = fill("2E75B6")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin()
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[5].height = 22

        # ── Data rows ───────────────────────────────────────────────────────
        for idx, r in enumerate(results, 1):
            row    = 5 + idx
            is_fail = r.status == "FAIL"
            bg = "FDEDEC" if is_fail else ("EAF7EE" if idx % 2 else "FFFFFF")
            ws.row_dimensions[row].height = 18

            def put(col, val, align="center", bold=False, color="000000",
                    wrap=False):
                c = ws.cell(row, col, val)
                c.font      = Font(name="Arial", size=10, bold=bold, color=color)
                c.fill      = fill(bg)
                c.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=wrap)
                c.border    = thin()

            put(1,  idx)
            put(2,  r.pin_name)
            put(3,  r.status,
                bold=is_fail, color="C0392B" if is_fail else "1E8449")
            put(4,  r.mi   if r.mi  >= 0 else "—")
            put(5,  r.ch   if r.ch  >= 0 else "—")
            put(6,  r.conn_count if r.status == "OK" else 0)
            put(7,  r.lat  if r.lat >= 0 else "—")
            put(8,  f"0x{r.out0:02X}" if r.out0 >= 0 else "—")
            put(9,  f"0x{r.out1:02X}" if r.out1 >= 0 else "—")
            put(10, r.root_cause,     "left", wrap=True)
            put(11, r.recommendation, "left", wrap=True)

        ws.freeze_panes = "A6"

    def _build_jig_sheet_fail(self, ws, failed, side_label="J1"):
        """Sheet 'Root Cause Analysis' — chỉ pin FAIL, trình bày dạng card."""
        thin = self._jig_thin
        fill = self._jig_fill

        ws.sheet_view.showGridLines = False
        for col, w in zip("ABC", [6, 22, 58]):
            ws.column_dimensions[col].width = w

        # Title
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value     = f"ROOT CAUSE ANALYSIS — {side_label} — {len(failed)} FAILED PIN(S)"
        c.font      = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        c.fill      = fill("922B21")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        row = 3
        for idx, r in enumerate(failed, 1):
            # Pin header bar
            ws.merge_cells(f"A{row}:C{row}")
            c = ws[f"A{row}"]
            c.value     = (f"#{idx}  {side_label}.{r.pin_name}"
                           f"   MUX group {r.mi},  channel {r.ch}")
            c.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            c.fill      = fill("C0392B")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 24
            row += 1

            drive_str = (f"LAT={r.lat}  "
                         f"OUT0=0x{r.out0:02X}  "
                         f"OUT1=0x{r.out1:02X}"
                         if r.lat >= 0 else "—")

            for label, value, lbl_fill, val_fill, h in [
                ("Drive State",  drive_str,         "FAD7A0", "FEF9E7", 18),
                ("Nguyên nhân",  r.root_cause,       "F5CBA7", "FEF9E7", 36),
                ("Khắc phục",    r.recommendation,   "A9DFBF", "EAFAF1", 36),
            ]:
                ws[f"A{row}"].border = thin()
                ws[f"B{row}"].value = label
                ws[f"B{row}"].font  = Font(name="Arial", size=10, bold=True)
                ws[f"B{row}"].fill  = fill(lbl_fill)
                ws[f"B{row}"].alignment = Alignment(
                    horizontal="left", vertical="top", indent=1)
                ws[f"B{row}"].border = thin()
                ws[f"C{row}"].value = value
                ws[f"C{row}"].font  = Font(name="Arial", size=10)
                ws[f"C{row}"].fill  = fill(val_fill)
                ws[f"C{row}"].alignment = Alignment(
                    horizontal="left", vertical="top",
                    wrap_text=True, indent=1)
                ws[f"C{row}"].border = thin()
                ws.row_dimensions[row].height = h
                row += 1

            row += 1   # blank separator between pins

    # ── Export ─────────────────────────────────────────────────────────────
    def _export_report(self):
        if not self.current_result:
            messagebox.showinfo("No Data", "No test result to export.")
            return

        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)

        ts    = self.current_result.timestamp.strftime("%Y%m%d_%H%M%S")
        sn    = self.current_result.serial_number.replace("/","_")
        fname = f"cable_test_{sn}_{ts}.xlsx"
        out   = reports_dir / fname

        template = Path(TEMPLATE_FILE)
        if not template.exists():
            # Auto-create template
            try:
                from create_template import create_template
                create_template(str(template))
            except Exception as e:
                messagebox.showerror("Template Error",
                    f"Cannot create template: {e}\n"
                    "Place cable_test_report_template.xlsx in app folder.")
                return

        try:
            gen = ReportGenerator()
            gen.generate(self.current_result, str(template), str(out))
            messagebox.showinfo("Report Saved",
                f"Report saved:\n{out}\n\nOpen now?",
            )
            if messagebox.askyesno("Open?", f"Open {fname}?"):
                os.startfile(str(out)) if sys.platform=="win32" else \
                os.system(f"xdg-open '{out}'")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _open_reports_folder(self):
        d = Path("reports")
        d.mkdir(exist_ok=True)
        if sys.platform == "win32":
            os.startfile(str(d))
        else:
            os.system(f"xdg-open '{d}'")

    # ── History ────────────────────────────────────────────────────────────
    def _save_history(self):
        try:
            data = []
            for r in self.history[-200:]:
                data.append({
                    "ts":          r.timestamp.isoformat(),
                    "sn":          r.serial_number,
                    "connections": len(r.connections),
                    "completed":   r.completed,
                    "duration":    r.duration_s,
                })
            with open(HISTORY_FILE, "w") as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def _load_history(self):
        try:
            if os.path.exists(HISTORY_FILE):
                with open(HISTORY_FILE) as f:
                    data = json.load(f)
                for d in data:
                    r = TestResult()
                    r.timestamp     = datetime.fromisoformat(d["ts"])
                    r.serial_number = d.get("sn", "")
                    r.completed     = d.get("completed", False)
                    r.duration_s    = d.get("duration", 0)
                    self.history.append(r)
                self._update_stats()
        except Exception:
            pass


# ── Entry point ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()